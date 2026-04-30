
import sys
# Force unbuffered output for Windows CMD with UTF-8
sys.stdout.reconfigure(line_buffering=True, encoding='utf-8', errors='replace')
sys.stderr.reconfigure(line_buffering=True, encoding='utf-8', errors='replace')

# Load .env FIRST before anything else reads env vars
from pathlib import Path as _Path
from dotenv import load_dotenv as _load_dotenv
_load_dotenv(_Path(__file__).resolve().parent / '.env', override=True)

from flask import Flask, request, jsonify, send_from_directory, send_file, Response
from flask_cors import CORS
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect, generate_csrf, validate_csrf, CSRFError
import hmac
from models import init_db, get_db, Country, Source, Keyword, Article, User, AuditLog, ExportRecord, UserFile, SearchHistory, BookmarkedArticle, DailyBrief
import uuid
from rss_service import fetch_feed
from translation_service import (
    translate_keyword, 
    detect_language, 
    translate_to_arabic
)
# New multilingual services
from keyword_expansion import expand_keyword, get_all_expansions, load_expansions_from_keywords
# multilingual_matcher imported where needed
# translation_cache imported where needed
# arabic_utils imported where needed
# multilingual_monitor - legacy, replaced by async_monitor_wrapper
# New optimized async services
from async_monitor_wrapper import run_optimized_monitoring, save_matched_articles_sync
# Utils
from utils import clean_html_content
from datetime import datetime, timedelta
from functools import wraps
import json
import os
import re
import threading as _threading

app = Flask(__name__, static_folder='static', static_url_path='')

# Ensure JSON responses use UTF-8 and do not escape Arabic characters
app.config['JSON_AS_ASCII'] = False
try:
    # Flask >= 2.3
    app.json.ensure_ascii = False  # type: ignore[attr-defined]
except Exception:
    pass

# ── Security config ──────────────────────────────────────────────────
_is_production = bool(os.getenv('RENDER') or os.getenv('FLASK_ENV') == 'production')

# C2 FIX: No hardcoded secret. Fail in production; use ephemeral in dev.
_secret = os.environ.get('SECRET_KEY', '')
if not _secret:
    if _is_production:
        raise RuntimeError('SECRET_KEY environment variable is required in production')
    import secrets as _sec
    _secret = _sec.token_urlsafe(32)
    print('[APP] ⚠️ SECRET_KEY not set — using ephemeral dev secret')
app.secret_key = _secret

app.config.setdefault('SESSION_COOKIE_HTTPONLY', True)
app.config.setdefault('SESSION_COOKIE_SAMESITE', 'Lax')
if _is_production or os.environ.get('SESSION_COOKIE_SECURE', 'false').lower() == 'true':
    app.config['SESSION_COOKIE_SECURE'] = True

# C3 FIX: Disable Flask-WTF auto-check; we enforce CSRF via before_request instead.
app.config['WTF_CSRF_CHECK_DEFAULT'] = False

login_manager = LoginManager(app)
csrf = CSRFProtect(app)

@login_manager.unauthorized_handler
def unauthorized_json():
    """Return JSON 401 instead of HTML so the SPA can handle it."""
    return jsonify({"error": "يرجى تسجيل الدخول أولاً", "code": "unauthorized"}), 401

# H6 FIX: Restrict CORS origins (env var or dev-only localhost list).
_cors_origins = os.environ.get('CORS_ORIGINS', '').strip()
if _cors_origins:
    _allowed_origins = [o.strip() for o in _cors_origins.split(',') if o.strip()]
else:
    _allowed_origins = [
        'http://localhost:5173',
        'http://localhost:5555',
        'http://127.0.0.1:5173',
        'http://127.0.0.1:5555',
    ]
CORS(app, supports_credentials=True, origins=_allowed_origins)

# Initialize database on startup
init_db()
try:
    from models import engine as _engine
    print(f"[DB] Connected to database: {_engine.url}")
except Exception:
    pass

# Database session management
@app.teardown_appcontext
def shutdown_session(exception=None):
    """Close database sessions after each request"""
    pass  # Sessions are now managed per-request

# C3 FIX: Custom CSRF enforcement for SPA — validates token on authenticated
# state-changing requests.  Login/signup/external-API are exempt (pre-auth).
_CSRF_EXEMPT_PREFIXES = ('/api/auth/login', '/api/auth/signup', '/api/auth/check', '/api/external/')

@app.before_request
def _csrf_protect():
    if request.method in ('GET', 'HEAD', 'OPTIONS', 'TRACE'):
        return
    if any(request.path.startswith(p) for p in _CSRF_EXEMPT_PREFIXES):
        return
    if not current_user.is_authenticated:
        return
    token = request.headers.get('X-CSRFToken') or request.headers.get('X-CSRF-Token')
    if not token:
        return jsonify({'error': 'Missing CSRF token'}), 403
    try:
        validate_csrf(token)
    except CSRFError:
        return jsonify({'error': 'Invalid CSRF token'}), 403

@app.after_request
def _set_csrf_cookie(response):
    if response.status_code < 400:
        try:
            token = generate_csrf()
            is_secure = app.config.get('SESSION_COOKIE_SECURE', False)
            response.set_cookie(
                'csrf_token', token,
                httponly=False,       # JS must read this
                samesite='Lax',
                secure=is_secure,
                path='/',
            )
        except Exception:
            pass
    return response


def admin_required(fn):
    """Decorator that requires the current user to be an ADMIN."""
    @wraps(fn)
    @login_required
    def wrapper(*args, **kwargs):
        try:
            print(f"[AUTH] admin_required check: auth={current_user.is_authenticated}, id={getattr(current_user, 'id', None)}, email={getattr(current_user, 'email', None)}, role={getattr(current_user, 'role', None)}")
        except Exception:
            pass
        if getattr(current_user, "role", "USER") != "ADMIN":
            return jsonify({"error": "Forbidden"}), 403
        return fn(*args, **kwargs)

    return wrapper


def scoped(query, Model, force_user_filter=False):
    """Scope queries to the current user for models that have user_id.

    - If force_user_filter=True: ALWAYS filter by current user (even for ADMIN)
    - ADMIN sees all rows ONLY if force_user_filter=False (for admin dashboards)
    - Regular users see only rows where Model.user_id == current_user.id.
    - Anonymous users get the original query (public/global data only).
    
    IMPORTANT: For user-owned data (articles, keywords, exports), always use
    force_user_filter=True to ensure proper isolation.
    """
    try:
        if not current_user.is_authenticated:
            return query
        
        # For user-owned data, always filter by user_id (even for admin)
        if force_user_filter:
            if hasattr(Model, "user_id"):
                return query.filter(Model.user_id == current_user.id)
            return query
        
        # Legacy behavior: ADMIN sees all (only for admin dashboards)
        if getattr(current_user, "role", "USER") == "ADMIN":
            return query
        
        if hasattr(Model, "user_id"):
            return query.filter(Model.user_id == current_user.id)
        return query
    except RuntimeError:
        # Outside request context; just return original query
        return query


def get_current_user_id():
    """Get current user ID safely, returns None if not authenticated."""
    try:
        if current_user.is_authenticated:
            return current_user.id
        return None
    except RuntimeError:
        return None


def log_action(user_id=None, admin_id=None, action="", meta=None):
    """Write a simple audit log entry.

    meta can be any JSON-serializable object; it will be stored as JSON text.
    Errors are swallowed so logging never breaks main flows.
    """
    try:
        db = get_db()
        try:
            meta_text = None
            if meta is not None:
                try:
                    meta_text = json.dumps(meta, ensure_ascii=False)
                except Exception:
                    meta_text = str(meta)

            entry = AuditLog(
                user_id=user_id,
                admin_id=admin_id,
                action=action or "",
                meta_json=meta_text,
            )
            db.add(entry)
            db.commit()
        finally:
            db.close()
    except Exception:
        # Never fail the main request because of audit logging
        pass


@app.route('/api/admin/audit', methods=['GET'])
@admin_required
def admin_audit_logs():
    """Return recent audit log entries for admins.

    Optional query params:
      - user_id: filter by user
      - limit: max number of rows (default 50)
    """
    user_id = request.args.get('user_id', type=int)
    limit = request.args.get('limit', default=50, type=int)
    limit = max(1, min(limit, 200))

    db = get_db()
    try:
        query = db.query(AuditLog).order_by(AuditLog.created_at.desc())
        if user_id:
            query = query.filter(AuditLog.user_id == user_id)
        logs = query.limit(limit).all()
        result = []
        for row in logs:
            try:
                meta = json.loads(row.meta_json) if row.meta_json else None
            except Exception:
                meta = row.meta_json
            result.append({
                'id': row.id,
                'user_id': row.user_id,
                'admin_id': row.admin_id,
                'action': row.action,
                'meta': meta,
                'created_at': row.created_at.isoformat() if row.created_at else None,
            })
        return jsonify(result)
    finally:
        db.close()


@app.route('/api/exports', methods=['GET'])
@login_required
def list_exports():
    """List export history. Admins see all files, users see only their own."""
    db = get_db()
    try:
        is_admin = current_user.role == 'ADMIN'
        
        if is_admin:
            # Admin sees all exports with user info
            exports = db.query(ExportRecord).order_by(ExportRecord.created_at.desc()).all()
        else:
            # Regular user sees only their exports
            exports = db.query(ExportRecord).filter(
                ExportRecord.user_id == current_user.id
            ).order_by(ExportRecord.created_at.desc()).all()
        
        # Build user lookup for admin view
        user_map = {}
        if is_admin:
            user_ids = set(rec.user_id for rec in exports)
            users = db.query(User).filter(User.id.in_(user_ids)).all()
            user_map = {u.id: u.name for u in users}
        
        result = []
        for rec in exports:
            try:
                filters = json.loads(rec.filters_json) if rec.filters_json else None
            except Exception:
                filters = None
            item = {
                'id': rec.id,
                'article_count': rec.article_count,
                'filters': filters,
                'filename': rec.filename,
                'file_size': rec.file_size,
                'has_file': bool(rec.stored_filename),
                'source_type': rec.source_type or 'dashboard',
                'created_at': rec.created_at.isoformat() if rec.created_at else None,
            }
            if is_admin:
                item['user_id'] = rec.user_id
                item['user_name'] = user_map.get(rec.user_id, 'مستخدم غير معروف')
            result.append(item)
        return jsonify(result)
    finally:
        db.close()


# Legacy filesystem storage (kept for backward compatibility)
EXPORTS_FOLDER = os.path.join(os.path.dirname(__file__), 'exports')
os.makedirs(EXPORTS_FOLDER, exist_ok=True)

def get_user_exports_folder(user_id):
    """Get or create user-specific exports folder (legacy)"""
    user_folder = os.path.join(EXPORTS_FOLDER, str(user_id))
    os.makedirs(user_folder, exist_ok=True)
    return user_folder


@app.route('/api/exports/generate-pdf', methods=['POST'])
@login_required
def generate_pdf():
    """Generate a real PDF from article data using reportlab (pure Python)."""
    data = request.get_json() or {}
    articles = data.get('articles', [])
    title = data.get('title', 'تقرير أخبار عين')
    stats = data.get('stats', None)
    brief = data.get('brief', None)

    if not articles:
        return jsonify({'error': 'No articles provided'}), 400

    try:
        from pdf_generator import generate_report_pdf
        pdf_bytes = generate_report_pdf(articles, title=title, stats=stats, brief=brief)
        return Response(
            pdf_bytes,
            mimetype='application/pdf',
            headers={'Content-Disposition': 'inline; filename="report.pdf"'}
        )
    except Exception as e:
        print(f"[PDF] ❌ PDF generation error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'PDF generation failed: {str(e)}'}), 500


@app.route('/api/exports', methods=['POST'])
@login_required
def create_export_record():
    """Create an export record for the current user with optional file upload.

    Accepts multipart form data with:
    - file: PDF file (optional)
    - filters: JSON string of filters
    - article_count: number of articles
    
    Files are stored in database (file_data column) for Render compatibility.
    """
    # Handle both JSON and multipart form data
    if request.content_type and 'multipart/form-data' in request.content_type:
        filters_str = request.form.get('filters', '{}')
        article_count = int(request.form.get('article_count', 0))
        source_type = request.form.get('source_type', 'dashboard')
        file = request.files.get('file')
        try:
            filters = json.loads(filters_str)
        except:
            filters = {}
    else:
        data = request.get_json() or {}
        filters = data.get('filters') or {}
        article_count = int(data.get('article_count') or 0)
        source_type = data.get('source_type', 'dashboard')
        file = None

    db = get_db()
    try:
        rec = ExportRecord(
            user_id=current_user.id,
            filters_json=json.dumps(filters, ensure_ascii=False) if filters else None,
            article_count=article_count,
            source_type=source_type,
        )
        
        # Handle file upload - store in database for Render compatibility
        if file and file.filename:
            file_data = file.read()
            rec.filename = file.filename
            rec.file_data = file_data
            rec.file_size = len(file_data)
            # stored_filename kept for legacy compatibility
            rec.stored_filename = f"{uuid.uuid4().hex}.pdf"
        
        db.add(rec)
        db.commit()
        try:
            log_action(user_id=current_user.id, action="export_pdf", meta={
                'export_id': rec.id,
                'article_count': article_count,
            })
        except Exception:
            pass
        return jsonify({'success': True, 'id': rec.id}), 201
    finally:
        db.close()


@app.route('/api/exports/<int:export_id>/download', methods=['GET'])
@login_required
def download_export(export_id):
    """Download an exported file. Admins can download any file.
    
    Query params:
    - view=1: View inline (for PDF preview)
    - (default): Download as attachment
    
    Files are retrieved from database (file_data column) for Render compatibility.
    Falls back to filesystem for legacy files.
    """
    db = get_db()
    try:
        is_admin = current_user.role == 'ADMIN'
        view_mode = request.args.get('view', '0') == '1'
        
        if is_admin:
            rec = db.query(ExportRecord).filter(ExportRecord.id == export_id).first()
        else:
            rec = db.query(ExportRecord).filter(
                ExportRecord.id == export_id,
                ExportRecord.user_id == current_user.id
            ).first()
        
        if not rec:
            return jsonify({'error': 'السجل غير موجود'}), 404
        
        # Detect mimetype from filename
        import mimetypes as _mt
        fname = rec.filename or f"export_{export_id}.pdf"
        guessed_mime = _mt.guess_type(fname)[0] or 'application/pdf'
        
        # Try database storage first (new method for Render)
        if rec.file_data:
            from io import BytesIO
            file_stream = BytesIO(rec.file_data)
            return send_file(
                file_stream,
                mimetype=guessed_mime,
                as_attachment=not view_mode,
                download_name=fname
            )
        
        # Fallback to filesystem (legacy files)
        if rec.stored_filename:
            user_folder = get_user_exports_folder(rec.user_id)
            file_path = os.path.join(user_folder, rec.stored_filename)
            
            if os.path.exists(file_path):
                return send_file(
                    file_path,
                    mimetype=guessed_mime,
                    as_attachment=not view_mode,
                    download_name=fname
                )
        
        return jsonify({'error': 'لا يوجد ملف مرفق'}), 404
    finally:
        db.close()


@app.route('/api/exports/<int:export_id>', methods=['DELETE'])
@login_required
def delete_export(export_id):
    """Delete an export record and its file. Admins can delete any file.
    
    Deletes from database. Also cleans up legacy filesystem files if present.
    """
    db = get_db()
    try:
        is_admin = current_user.role == 'ADMIN'
        
        if is_admin:
            rec = db.query(ExportRecord).filter(ExportRecord.id == export_id).first()
        else:
            rec = db.query(ExportRecord).filter(
                ExportRecord.id == export_id,
                ExportRecord.user_id == current_user.id
            ).first()
        
        if not rec:
            return jsonify({'error': 'السجل غير موجود'}), 404
        
        # Delete legacy filesystem file if exists
        if rec.stored_filename and not rec.file_data:
            user_folder = get_user_exports_folder(rec.user_id)
            file_path = os.path.join(user_folder, rec.stored_filename)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    pass
        
        db.delete(rec)
        db.commit()
        return jsonify({'success': True})
    finally:
        db.close()


# ==================== User Files APIs ====================
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls', 'csv', 'doc', 'docx', 'txt', 'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_user_upload_folder(user_id):
    """Get or create user-specific upload folder"""
    user_folder = os.path.join(UPLOAD_FOLDER, str(user_id))
    os.makedirs(user_folder, exist_ok=True)
    return user_folder


@app.route('/api/files', methods=['GET'])
@login_required
def list_user_files():
    """List files for the current user."""
    db = get_db()
    try:
        files = db.query(UserFile).filter(
            UserFile.user_id == current_user.id
        ).order_by(UserFile.created_at.desc()).all()
        result = []
        for f in files:
            result.append({
                'id': f.id,
                'filename': f.filename,
                'file_type': f.file_type,
                'file_size': f.file_size,
                'description': f.description,
                'created_at': f.created_at.isoformat() if f.created_at else None,
            })
        return jsonify(result)
    finally:
        db.close()


@app.route('/api/files', methods=['POST'])
@login_required
def upload_user_file():
    """Upload a file for the current user."""
    if 'file' not in request.files:
        return jsonify({'error': 'لم يتم تحديد ملف'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'لم يتم تحديد ملف'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'نوع الملف غير مسموح'}), 400
    
    # Generate unique stored filename
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    stored_filename = f"{uuid.uuid4().hex}.{ext}" if ext else uuid.uuid4().hex
    
    # Save file to user folder
    user_folder = get_user_upload_folder(current_user.id)
    file_path = os.path.join(user_folder, stored_filename)
    file.save(file_path)
    
    # Get file size
    file_size = os.path.getsize(file_path)
    
    # Get description from form data
    description = request.form.get('description', '')
    
    db = get_db()
    try:
        user_file = UserFile(
            user_id=current_user.id,
            filename=file.filename,
            stored_filename=stored_filename,
            file_type=ext,
            file_size=file_size,
            description=description,
        )
        db.add(user_file)
        db.commit()
        return jsonify({
            'success': True,
            'id': user_file.id,
            'filename': user_file.filename,
        }), 201
    finally:
        db.close()


@app.route('/api/files/<int:file_id>', methods=['GET'])
@login_required
def download_user_file(file_id):
    """Download a user file."""
    db = get_db()
    try:
        user_file = db.query(UserFile).filter(
            UserFile.id == file_id,
            UserFile.user_id == current_user.id
        ).first()
        if not user_file:
            return jsonify({'error': 'الملف غير موجود'}), 404
        
        user_folder = get_user_upload_folder(current_user.id)
        file_path = os.path.join(user_folder, user_file.stored_filename)
        
        if not os.path.exists(file_path):
            return jsonify({'error': 'الملف غير موجود على الخادم'}), 404
        
        from flask import send_file
        return send_file(
            file_path,
            download_name=user_file.filename,
            as_attachment=True
        )
    finally:
        db.close()


@app.route('/api/files/<int:file_id>', methods=['DELETE'])
@login_required
def delete_user_file(file_id):
    """Delete a user file."""
    db = get_db()
    try:
        user_file = db.query(UserFile).filter(
            UserFile.id == file_id,
            UserFile.user_id == current_user.id
        ).first()
        if not user_file:
            return jsonify({'error': 'الملف غير موجود'}), 404
        
        # Delete physical file
        user_folder = get_user_upload_folder(current_user.id)
        file_path = os.path.join(user_folder, user_file.stored_filename)
        if os.path.exists(file_path):
            os.remove(file_path)
        
        # Delete database record
        db.delete(user_file)
        db.commit()
        return jsonify({'success': True})
    finally:
        db.close()


# ==================== Search History APIs ====================

@app.route('/api/search-history', methods=['GET'])
@login_required
def list_search_history():
    """List search history for the current user."""
    db = get_db()
    try:
        history = db.query(SearchHistory).filter(
            SearchHistory.user_id == current_user.id
        ).order_by(SearchHistory.created_at.desc()).limit(50).all()
        result = []
        for h in history:
            try:
                filters = json.loads(h.filters_json) if h.filters_json else None
            except Exception:
                filters = None
            result.append({
                'id': h.id,
                'search_type': h.search_type,
                'query': h.query,
                'filters': filters,
                'results_count': h.results_count,
                'created_at': h.created_at.isoformat() if h.created_at else None,
            })
        return jsonify(result)
    finally:
        db.close()


@app.route('/api/search-history', methods=['POST'])
@login_required
def create_search_history():
    """Record a search in history."""
    data = request.get_json() or {}
    search_type = data.get('search_type', 'keyword')
    query = data.get('query', '')
    filters = data.get('filters') or {}
    results_count = int(data.get('results_count') or 0)

    db = get_db()
    try:
        rec = SearchHistory(
            user_id=current_user.id,
            search_type=search_type,
            query=query,
            filters_json=json.dumps(filters, ensure_ascii=False) if filters else None,
            results_count=results_count,
        )
        db.add(rec)
        db.commit()
        return jsonify({'success': True, 'id': rec.id}), 201
    finally:
        db.close()


@app.route('/api/search-history/<int:history_id>', methods=['DELETE'])
@login_required
def delete_search_history(history_id):
    """Delete a search history entry."""
    db = get_db()
    try:
        rec = db.query(SearchHistory).filter(
            SearchHistory.id == history_id,
            SearchHistory.user_id == current_user.id
        ).first()
        if not rec:
            return jsonify({'error': 'السجل غير موجود'}), 404
        db.delete(rec)
        db.commit()
        return jsonify({'success': True})
    finally:
        db.close()


# ==================== Admin APIs ====================

@app.route('/api/admin/users', methods=['GET'])
@admin_required
def admin_list_users():
    """List users with optional search by name/email (ADMIN only)."""
    q = (request.args.get('q') or '').strip().lower()
    db = get_db()
    try:
        query = db.query(User)
        if q:
            like = f"%{q}%"
            query = query.filter(
                (User.email.ilike(like)) | (User.name.ilike(like))
            )
        users = query.order_by(User.created_at.desc()).all()
        result = []
        for u in users:
            keyword_count = db.query(Keyword).filter(Keyword.user_id == u.id, Keyword.enabled == True).count()
            article_count = db.query(Article).filter(Article.user_id == u.id).count()
            result.append({
                'id': u.id,
                'name': u.name,
                'email': u.email,
                'role': u.role,
                'is_active': u.is_active,
                'created_at': u.created_at.isoformat() if u.created_at else None,
                'keyword_count': keyword_count,
                'article_count': article_count,
            })
        return jsonify(result)
    finally:
        db.close()


@app.route('/api/admin/users', methods=['POST'])
@admin_required
def admin_create_user():
    """Create a new user (ADMIN only).

    Expects JSON: {name, email, role?, is_active?, password?}
    """
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()
    name = (data.get('name') or '').strip()
    role = data.get('role') or 'USER'
    is_active = bool(data.get('is_active', True))
    # H1 FIX: Generate secure random password if none provided (never default to '0000')
    import secrets as _s
    password = data.get('password') or _s.token_urlsafe(12)

    if not email or not name:
        return jsonify({'error': 'Name and email are required'}), 400

    if role not in ('ADMIN', 'USER'):
        role = 'USER'

    db = get_db()
    try:
        existing = db.query(User).filter(User.email == email).first()
        if existing:
            return jsonify({'error': 'User with this email already exists'}), 400

        from auth_utils import hash_password
        user = User(
            email=email,
            name=name,
            role=role,
            is_active=is_active,
            password_hash=hash_password(password),
        )
        db.add(user)
        db.commit()
        try:
            log_action(
                admin_id=getattr(current_user, 'id', None),
                user_id=user.id,
                action="admin_create_user",
                meta={"email": email, "role": role}
            )
        except Exception:
            pass

        return jsonify({'success': True, 'id': user.id}), 201
    finally:
        db.close()


@app.route('/api/admin/users/<int:user_id>', methods=['PATCH'])
@admin_required
def admin_update_user(user_id):
    """Update basic user fields (name, role, is_active, optional password reset)."""
    data = request.get_json() or {}
    db = get_db()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return jsonify({'error': 'User not found'}), 404

        if 'name' in data:
            user.name = data['name']
        if 'email' in data:
            new_email = (data.get('email') or '').strip().lower()
            if new_email:
                user.email = new_email
        if 'role' in data and data['role'] in ('ADMIN', 'USER'):
            user.role = data['role']
        if 'is_active' in data:
            user.is_active = bool(data['is_active'])
        from auth_utils import hash_password
        if data.get('reset_password'):
            # H1 FIX: Generate random temp password instead of weak '0000'
            import secrets as _s
            _temp_pw = _s.token_urlsafe(12)
            user.password_hash = hash_password(_temp_pw)
            user.must_change_password = True
        if data.get('password'):
            user.password_hash = hash_password(data['password'])
        db.commit()
        try:
            log_action(
                admin_id=getattr(current_user, 'id', None),
                user_id=user.id,
                action="admin_update_user",
                meta={k: data[k] for k in data.keys() if k != 'password'}
            )
        except Exception:
            pass
        return jsonify({'success': True})
    finally:
        db.close()


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@admin_required
def admin_delete_user(user_id):
    """Delete a user (ADMIN only)."""
    db = get_db()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return jsonify({'error': 'User not found'}), 404

        # Prevent admin from deleting themselves
        try:
            if current_user.is_authenticated and current_user.id == user.id:
                return jsonify({'error': 'Cannot delete currently logged-in user'}), 400
        except Exception:
            pass

        # Cascade-delete all related data (PostgreSQL enforces FK constraints)
        from models import Article, Keyword, AuditLog, ExportRecord, UserFile, SearchHistory, MonitorJob
        try:
            from models import UserArticle, UserCountry, UserSource
            db.query(UserArticle).filter(UserArticle.user_id == user_id).delete()
            db.query(UserCountry).filter(UserCountry.user_id == user_id).delete()
            db.query(UserSource).filter(UserSource.user_id == user_id).delete()
        except Exception:
            pass
        db.query(Article).filter(Article.user_id == user_id).delete()
        db.query(Keyword).filter(Keyword.user_id == user_id).delete()
        db.query(ExportRecord).filter(ExportRecord.user_id == user_id).delete()
        db.query(MonitorJob).filter(MonitorJob.user_id == user_id).delete()
        db.query(SearchHistory).filter(SearchHistory.user_id == user_id).delete()
        db.query(UserFile).filter(UserFile.user_id == user_id).delete()
        db.query(AuditLog).filter(AuditLog.user_id == user_id).delete()
        db.query(AuditLog).filter(AuditLog.admin_id == user_id).update({AuditLog.admin_id: None})

        user_email = user.email
        db.delete(user)
        db.commit()
        try:
            log_action(
                admin_id=getattr(current_user, 'id', None),
                user_id=None,
                action="admin_delete_user",
                meta={"email": user_email}
            )
        except Exception:
            pass
        return jsonify({'success': True})
    finally:
        db.close()


@app.route('/api/admin/users/<int:user_id>/keywords', methods=['GET'])
@admin_required
def admin_get_user_keywords(user_id):
    """Get all keywords for a specific user (ADMIN only)."""
    db = get_db()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return jsonify({'error': 'User not found'}), 404

        keywords = db.query(Keyword).filter(Keyword.user_id == user_id).order_by(Keyword.created_at.desc()).all()

        result = []
        for kw in keywords:
            # Count articles matched by this keyword for this user
            article_count = db.query(Article).filter(
                Article.user_id == user_id,
                Article.keyword_original == kw.text_ar
            ).count()

            has_translations = bool(kw.translations_json)
            translations_age = None
            if kw.translations_updated_at:
                age = datetime.utcnow() - kw.translations_updated_at
                translations_age = age.days

            result.append({
                'id': kw.id,
                'text_ar': kw.text_ar,
                'text_en': kw.text_en,
                'enabled': kw.enabled,
                'has_translations': has_translations,
                'translations_age_days': translations_age,
                'article_count': article_count,
                'created_at': kw.created_at.isoformat() if kw.created_at else None,
            })

        return jsonify({
            'user_id': user_id,
            'user_name': user.name or user.email,
            'keywords': result,
            'total': len(result),
        })
    finally:
        db.close()


@app.route('/api/admin/stats', methods=['GET'])
@admin_required
def admin_stats():
    """Basic system statistics for admin dashboard."""
    db = get_db()
    try:
        total_users = db.query(User).count()
        active_users = db.query(User).filter(User.is_active == True).count()
        total_keywords = db.query(Keyword).count()
        total_articles = db.query(Article).count()
        total_exports = db.query(ExportRecord).count()
        return jsonify({
            'total_users': total_users,
            'active_users': active_users,
            'total_keywords': total_keywords,
            'total_articles': total_articles,
            'total_exports': total_exports,
        })
    finally:
        db.close()


@app.route('/api/home/stats', methods=['GET'])
@login_required
def home_stats():
    """Aggregated dashboard stats for the home page (current user only)."""
    db = get_db()
    try:
        from sqlalchemy import func, or_, desc
        uid = current_user.id

        base = db.query(Article).filter(Article.user_id == uid)

        # Total articles
        total_articles = base.count()

        # Sentiment breakdown
        positive = base.filter(or_(Article.sentiment_label == 'إيجابي', Article.sentiment == 'إيجابي')).count()
        negative = base.filter(or_(Article.sentiment_label == 'سلبي', Article.sentiment == 'سلبي')).count()
        neutral = base.filter(or_(Article.sentiment_label == 'محايد', Article.sentiment == 'محايد')).count()

        # Articles per country
        countries_q = db.query(
            Article.country, func.count(Article.id).label('cnt')
        ).filter(Article.user_id == uid, Article.country.isnot(None), Article.country != '').group_by(Article.country).order_by(desc('cnt')).all()
        countries_data = [{'name': c, 'count': n} for c, n in countries_q]

        # Top keywords by frequency (keyword_original field)
        kw_q = db.query(
            Article.keyword_original, func.count(Article.id).label('cnt')
        ).filter(Article.user_id == uid, Article.keyword_original.isnot(None), Article.keyword_original != '').group_by(Article.keyword_original).order_by(desc('cnt')).limit(10).all()
        top_keywords = [{'keyword': k, 'count': n} for k, n in kw_q]

        # User keyword count
        kw_count = db.query(Keyword).filter(Keyword.user_id == uid).count()

        # Unique countries with results
        unique_countries = db.query(Article.country).filter(Article.user_id == uid).distinct().count()

        # Bookmarks count
        bookmark_count = 0
        try:
            from models import Bookmark as BM
            bookmark_count = db.query(BM).filter(BM.user_id == uid).count()
        except Exception:
            pass

        return jsonify({
            'total_articles': total_articles,
            'positive': positive,
            'negative': negative,
            'neutral': neutral,
            'unique_countries': unique_countries,
            'keyword_count': kw_count,
            'bookmark_count': bookmark_count,
            'countries': countries_data,
            'top_keywords': top_keywords,
        })
    finally:
        db.close()


# ── System-wide map data (all users aggregated, cached 30 min) ──────
_map_cache = {'data': None, 'ts': None, 'lock': _threading.Lock()}
_MAP_CACHE_TTL = 1800  # 30 minutes in seconds


def _build_map_cache():
    """Build aggregated map data from ALL users' articles.
    
    Optimised: ~6 queries total instead of 30+ (eliminated N+1 country loop).
    """
    db = get_db()
    try:
        from sqlalchemy import func, or_, desc, case, text as sa_text

        # ── 1. Dedup base: one row per unique URL ───────────────────────
        dedup_sub = db.query(
            func.min(Article.id).label('id')
        ).group_by(Article.url).subquery()

        # ── 2. Total + sentiment counts in ONE query ────────────────────
        sent_col = func.coalesce(Article.sentiment_label, Article.sentiment)
        totals = db.query(
            func.count(Article.id),
            func.sum(case((sent_col == 'إيجابي', 1), else_=0)),
            func.sum(case((sent_col == 'سلبي', 1), else_=0)),
            func.sum(case((sent_col == 'محايد', 1), else_=0)),
        ).join(dedup_sub, Article.id == dedup_sub.c.id).first()
        total_articles = totals[0] or 0
        positive = int(totals[1] or 0)
        negative = int(totals[2] or 0)
        neutral  = int(totals[3] or 0)

        # ── 3. Articles per country (distinct URLs) ─────────────────────
        countries_q = db.query(
            Article.country, func.count(func.distinct(Article.url)).label('cnt')
        ).filter(
            Article.country.isnot(None), Article.country != ''
        ).group_by(Article.country).order_by(desc('cnt')).all()
        countries_data = [{'name': c, 'count': n} for c, n in countries_q]

        # ── 4. Top keywords (distinct URLs) ─────────────────────────────
        kw_q = db.query(
            Article.keyword_original, func.count(func.distinct(Article.url)).label('cnt')
        ).filter(
            Article.keyword_original.isnot(None), Article.keyword_original != ''
        ).group_by(Article.keyword_original).order_by(desc('cnt')).limit(15).all()
        top_keywords = [{'keyword': k, 'count': n} for k, n in kw_q]

        # ── 5. Unique countries + active users ──────────────────────────
        unique_countries = len(countries_data)
        active_users = db.query(Article.user_id).distinct().count()

        # ── 6. Top sources (distinct URLs) ──────────────────────────────
        src_q = db.query(
            Article.source_name, func.count(func.distinct(Article.url)).label('cnt')
        ).filter(
            Article.source_name.isnot(None), Article.source_name != ''
        ).group_by(Article.source_name).order_by(desc('cnt')).limit(10).all()
        top_sources = [{'name': s, 'count': n} for s, n in src_q]

        # ── 7. Top articles per country — SINGLE query (no N+1 loop) ───
        # Pick newest row per URL, then rank by country, take top 8 each
        newest_per_url = db.query(
            func.max(Article.id).label('id'),
            Article.country.label('country'),
        ).filter(
            Article.country.isnot(None), Article.country != ''
        ).group_by(Article.url, Article.country).subquery()

        ranked = db.query(
            Article,
            newest_per_url.c.country.label('c_name'),
        ).join(
            newest_per_url, Article.id == newest_per_url.c.id
        ).order_by(newest_per_url.c.country, Article.created_at.desc()).all()

        # Group in Python — fast since data is already fetched
        from collections import defaultdict
        top_country_articles = defaultdict(list)
        for art, c_name in ranked:
            if len(top_country_articles[c_name]) >= 8:
                continue
            top_country_articles[c_name].append({
                'id': art.id,
                'title_ar': art.title_ar,
                'summary_ar': art.summary_ar,
                'source_name': art.source_name,
                'sentiment': art.sentiment_label or art.sentiment,
                'url': art.url,
                'image_url': art.image_url,
                'keyword': art.keyword_original,
                'published_at': art.published_at.isoformat() if art.published_at else None,
            })

        return {
            'total_articles': total_articles,
            'positive': positive,
            'negative': negative,
            'neutral': neutral,
            'unique_countries': unique_countries,
            'active_users': active_users,
            'countries': countries_data,
            'top_keywords': top_keywords,
            'top_sources': top_sources,
            'country_articles': dict(top_country_articles),
            'last_refresh': datetime.utcnow().isoformat(),
        }
    finally:
        db.close()


@app.route('/api/home/map-data', methods=['GET'])
@login_required
def home_map_data():
    """System-wide aggregated map data. Cached for 30 minutes."""
    now = datetime.utcnow()
    force = request.args.get('force') == '1'

    with _map_cache['lock']:
        if (force or _map_cache['data'] is None or _map_cache['ts'] is None
                or (now - _map_cache['ts']).total_seconds() > _MAP_CACHE_TTL):
            _map_cache['data'] = _build_map_cache()
            _map_cache['ts'] = now

    return jsonify(_map_cache['data'])


@app.route('/api/home/map-timeline', methods=['GET'])
@login_required
def home_map_timeline():
    """Return article counts per country per day for the last N days (default 30).
    Response: { days: ['2026-04-01', ...], data: { 'السعودية': [0, 3, 5, ...], ... } }
    """
    from sqlalchemy import func, cast, Date as SADate
    days = min(int(request.args.get('days', 30)), 90)
    db = get_db()
    try:
        cutoff = datetime.utcnow() - timedelta(days=days)
        rows = db.query(
            cast(Article.created_at, SADate).label('day'),
            Article.country,
            func.count(func.distinct(Article.url)).label('cnt'),
        ).filter(
            Article.created_at >= cutoff,
            Article.country.isnot(None),
            Article.country != '',
        ).group_by('day', Article.country).all()

        # Build date range
        from datetime import date as _date
        start = (datetime.utcnow() - timedelta(days=days - 1)).date()
        all_days = [str(start + timedelta(days=i)) for i in range(days)]

        day_idx = {d: i for i, d in enumerate(all_days)}
        data = {}
        for day_val, country, cnt in rows:
            d_str = str(day_val) if not isinstance(day_val, str) else day_val
            if d_str not in day_idx:
                continue
            if country not in data:
                data[country] = [0] * len(all_days)
            data[country][day_idx[d_str]] = cnt

        return jsonify({'days': all_days, 'data': data})
    finally:
        db.close()


@app.route('/api/articles/countries/<path:country_name>/articles', methods=['GET'])
@login_required
def get_country_articles_system(country_name):
    """Get latest articles for a specific country (system-wide, deduplicated by URL)."""
    from sqlalchemy import func, desc
    db = get_db()
    try:
        # Pick the newest row for each distinct URL
        url_dedup = db.query(
            func.max(Article.id).label('id')
        ).filter(
            Article.country == country_name
        ).group_by(Article.url).order_by(desc('id')).limit(10).subquery()
        articles = db.query(Article).join(
            url_dedup, Article.id == url_dedup.c.id
        ).order_by(Article.created_at.desc()).all()
        return jsonify([{
            'id': a.id,
            'title_ar': a.title_ar,
            'summary_ar': a.summary_ar,
            'source_name': a.source_name,
            'sentiment': a.sentiment_label or a.sentiment,
            'url': a.url,
            'image_url': a.image_url,
            'keyword': a.keyword_original,
            'published_at': a.published_at.isoformat() if a.published_at else None,
        } for a in articles])
    finally:
        db.close()


# Whitelist for authentication (legacy; kept for compatibility)
ALLOWED_EMAILS = ["t09301970@gmail.com"]


@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    try:
        return db.get(User, int(user_id))
    finally:
        db.close()


@app.route('/')
def index():
    """Serve built frontend from the static folder (backend/static)."""
    return send_from_directory('static', 'index.html')

@app.route('/health')
def health_check():
    """Health check endpoint for Render"""
    return jsonify({'status': 'healthy', 'timestamp': datetime.utcnow().isoformat()}), 200

@app.route('/api/auth/check', methods=['POST'])
def check_auth():
    """Legacy auth check by email whitelist (kept for backward compatibility)."""
    data = request.get_json()
    email = data.get('email', '')
    
    if email in ALLOWED_EMAILS:
        return jsonify({"authorized": True})
    else:
        return jsonify({"authorized": False}), 403


# ==================== Session-based Auth ====================

from auth_utils import hash_password, verify_password


@app.route('/api/auth/signup', methods=['POST'])
def signup():
    """Public signup endpoint.

    Creates an inactive USER that must be approved by an admin (is_active=False).
    Expects JSON: {"name": str, "email": str, "password": str}
    """
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''

    if not name or not email or not password:
        return jsonify({"error": "Name, email and password are required"}), 400

    # H2 FIX: Enforce password complexity on signup
    from auth_utils import validate_password_strength
    pw_error = validate_password_strength(password)
    if pw_error:
        return jsonify({"error": pw_error}), 400

    db = get_db()
    try:
        existing = db.query(User).filter(User.email == email).first()
        if existing:
            return jsonify({"error": "User with this email already exists"}), 400

        user = User(
            name=name,
            email=email,
            password_hash=hash_password(password),
            role='USER',
            is_active=False,
            must_change_password=False,
        )
        db.add(user)
        db.commit()
        try:
            log_action(user_id=user.id, action="signup_requested", meta={"email": email})
        except Exception:
            pass

        return jsonify({"success": True, "pending_approval": True}), 201
    finally:
        db.close()


@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''

    if not email or not password:
        return jsonify({"error": "Email and password are required"}), 400

    db = get_db()
    try:
        user = db.query(User).filter(User.email == email).first()
        if not user or not verify_password(password, user.password_hash):
            return jsonify({"error": "بيانات الدخول غير صحيحة"}), 401
        if not user.is_active:
            return jsonify({"error": "حسابك غير مفعل. يرجى انتظار موافقة الإدارة"}), 403
        login_user(user)
        try:
            print(f"[AUTH] login success: id={user.id}, email={user.email}, role={user.role}")
        except Exception:
            pass
        try:
            log_action(user_id=user.id, action="login")
        except Exception:
            pass
        return jsonify({
            "id": user.id,
            "name": user.name,
            "email": user.email,
            "role": user.role
        })
    finally:
        db.close()


@app.route('/api/auth/profile', methods=['PATCH'])
@login_required
def update_profile():
    """Update current user's display name."""
    data = request.get_json() or {}
    new_name = (data.get('name') or '').strip()
    if not new_name:
        return jsonify({"error": "الاسم مطلوب"}), 400
    if len(new_name) > 100:
        return jsonify({"error": "الاسم طويل جداً"}), 400

    db = get_db()
    try:
        user = db.query(User).filter(User.id == current_user.id).first()
        if not user:
            return jsonify({"error": "User not found"}), 404
        user.name = new_name
        db.commit()
        return jsonify({
            "success": True,
            "name": user.name,
        })
    finally:
        db.close()


@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def change_password():
    """Change current user's password. Requires old password verification."""
    data = request.get_json() or {}
    old_password = data.get('old_password', '')
    new_password = data.get('new_password', '')

    if not old_password or not new_password:
        return jsonify({"error": "كلمة المرور القديمة والجديدة مطلوبتان"}), 400

    from auth_utils import validate_password_strength
    pw_error = validate_password_strength(new_password)
    if pw_error:
        return jsonify({"error": pw_error}), 400

    db = get_db()
    try:
        user = db.query(User).filter(User.id == current_user.id).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        if not verify_password(old_password, user.password_hash):
            return jsonify({"error": "كلمة المرور القديمة غير صحيحة"}), 400

        user.password_hash = hash_password(new_password)
        user.must_change_password = False
        db.commit()
        return jsonify({"success": True})
    finally:
        db.close()


@app.route('/api/auth/logout', methods=['POST'])
@login_required
def logout():
    uid = getattr(current_user, 'id', None)
    logout_user()
    try:
        if uid:
            log_action(user_id=uid, action="logout")
    except Exception:
        pass
    return jsonify({"success": True})


@app.route('/api/auth/me', methods=['GET'])
def auth_me():
    try:
        print(f"[AUTH] /api/auth/me: auth={current_user.is_authenticated}, id={getattr(current_user, 'id', None)}, email={getattr(current_user, 'email', None)}, role={getattr(current_user, 'role', None)}")
    except Exception:
        pass
    if not current_user.is_authenticated:
        return jsonify({"error": "Unauthorized"}), 401
    return jsonify({
        "id": current_user.id,
        "name": current_user.name,
        "email": current_user.email,
        "role": current_user.role
    })

# ==================== Countries ====================

@app.route('/api/countries', methods=['GET'])
@login_required
def get_countries():
    """Get all countries"""
    db = get_db()
    try:
        countries = db.query(Country).all()
        
        result = [{
            'id': c.id,
            'name_ar': c.name_ar,
            'enabled': c.enabled
        } for c in countries]
        
        return jsonify(result)
    finally:
        db.close()

@app.route('/api/countries/<int:country_id>/toggle', methods=['POST'])
@admin_required
def toggle_country(country_id):
    """Toggle country enabled status"""
    db = get_db()
    try:
        country = db.query(Country).filter(Country.id == country_id).first()
        
        if not country:
            return jsonify({"error": "Country not found"}), 404
        
        country.enabled = not country.enabled
        db.commit()
        
        return jsonify({"success": True, "enabled": country.enabled})
    finally:
        db.close()

# ==================== Sources ====================

@app.route('/api/sources', methods=['GET'])
@login_required
def get_sources():
    """Get all sources"""
    db = get_db()
    try:
        sources = db.query(Source).all()
        
        result = [{
            'id': s.id,
            'country_id': s.country_id,
            'country_name': s.country_name,
            'name': s.name,
            'url': s.url,
            'enabled': s.enabled,
            'fail_count': s.fail_count
        } for s in sources]
        
        return jsonify(result)
    finally:
        db.close()

@app.route('/api/sources', methods=['POST'])
@admin_required
def add_source():
    """Add new RSS source"""
    data = request.get_json()
    db = get_db()
    try:
        # Check if URL already exists
        existing = db.query(Source).filter(Source.url == data['url']).first()
        if existing:
            return jsonify({"error": "Source already exists"}), 400
        
        source = Source(
            country_id=data.get('country_id', 0),
            country_name=data['country_name'],
            name=data['name'],
            url=data['url'],
            enabled=True
        )
        
        db.add(source)
        db.commit()
        
        return jsonify({"success": True, "id": source.id})
    finally:
        db.close()

@app.route('/api/sources/<int:source_id>', methods=['PUT'])
@admin_required
def update_source(source_id):
    """Update source"""
    db = get_db()
    try:
        source = db.query(Source).filter(Source.id == source_id).first()
        
        if not source:
            return jsonify({"error": "Source not found"}), 404
        
        data = request.get_json()
        
        # Update fields
        if 'name' in data:
            source.name = data['name']
        if 'url' in data:
            source.url = data['url']
        if 'enabled' in data:
            source.enabled = data['enabled']
        
        db.commit()
        
        return jsonify({"success": True})
    finally:
        db.close()

@app.route('/api/sources/<int:source_id>', methods=['DELETE'])
@admin_required
def delete_source(source_id):
    """Delete source"""
    db = get_db()
    try:
        source = db.query(Source).filter(Source.id == source_id).first()
        
        if not source:
            return jsonify({"error": "Source not found"}), 404
        
        db.delete(source)
        db.commit()
        
        return jsonify({"success": True})
    finally:
        db.close()

@app.route('/api/sources/<int:source_id>/toggle', methods=['POST'])
@admin_required
def toggle_source(source_id):
    """Toggle source enabled status"""
    db = get_db()
    try:
        source = db.query(Source).filter(Source.id == source_id).first()
        
        if not source:
            return jsonify({"error": "Source not found"}), 404
        
        source.enabled = not source.enabled
        db.commit()
        
        return jsonify({"success": True, "enabled": source.enabled})
    finally:
        db.close()

# ==================== Keywords ====================

@app.route('/api/keywords', methods=['GET'])
@login_required
def get_keywords():
    """Get current user's keywords with translations.
    
    SECURITY: Uses force_user_filter=True to ensure each user
    only sees their own keywords, even admins.
    """
    db = get_db()
    try:
        # SECURITY FIX: Force user filter for proper isolation
        query = scoped(db.query(Keyword), Keyword, force_user_filter=True)
        keywords = query.all()
        
        result = [{
            'id': k.id,
            'text_ar': k.text_ar,
            'text_en': k.text_en,
            'text_fr': k.text_fr,
            'text_tr': k.text_tr,
            'text_ur': k.text_ur,
            'text_zh': k.text_zh,
            'text_ru': k.text_ru,
            'text_es': k.text_es,
            'enabled': k.enabled
        } for k in keywords]
        
        return jsonify(result)
    finally:
        db.close()

def copy_articles_from_shared_keyword(db, keyword_ar: str, target_user_id: int) -> int:
    """
    Copy articles from other users who have the same keyword.
    This allows sharing results between users with matching keywords.
    Returns the number of articles copied.
    """
    # Find other users who have this exact keyword
    other_keywords = db.query(Keyword).filter(
        Keyword.text_ar == keyword_ar,
        Keyword.user_id != target_user_id
    ).all()
    
    if not other_keywords:
        return 0
    
    # Get user IDs who have this keyword
    source_user_ids = [k.user_id for k in other_keywords]
    
    # Find articles from those users with this keyword
    source_articles = db.query(Article).filter(
        Article.user_id.in_(source_user_ids),
        Article.keyword == keyword_ar
    ).all()
    
    if not source_articles:
        return 0
    
    # Copy articles to target user (avoid duplicates by URL)
    existing_urls = set(
        a.url for a in db.query(Article).filter(
            Article.user_id == target_user_id
        ).all()
    )
    
    copied = 0
    for article in source_articles:
        if article.url in existing_urls:
            continue
        
        # Create a copy for the target user
        new_article = Article(
            country=article.country,
            source_name=article.source_name,
            url=article.url,
            title_original=article.title_original,
            summary_original=article.summary_original,
            original_language=article.original_language,
            image_url=article.image_url,
            title_ar=article.title_ar,
            summary_ar=article.summary_ar,
            arabic_text=article.arabic_text,
            keyword=article.keyword,
            keyword_original=article.keyword_original,
            keywords_translations=article.keywords_translations,
            sentiment_label=article.sentiment_label,
            sentiment_score=article.sentiment_score,
            published_at=article.published_at,
            fetched_at=article.fetched_at,
            user_id=target_user_id  # Assign to new user
        )
        db.add(new_article)
        existing_urls.add(article.url)
        copied += 1
    
    if copied > 0:
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[COPY] ⚠️ Error copying shared articles: {str(e)[:100]}")
            copied = 0
    
    return copied

@app.route('/api/keywords/expanded', methods=['GET'])
@login_required
def get_keywords_expanded():
    """Get current user's keywords with their multilingual expansions"""
    db = get_db()
    try:
        # SECURITY FIX: Get ONLY current user's enabled keywords
        keywords = db.query(Keyword).filter(
            Keyword.enabled == True,
            Keyword.user_id == current_user.id
        ).all()
        
        # Load/generate expansions for user's keywords only
        expansions = load_expansions_from_keywords(keywords)
        
        return jsonify(expansions)
    finally:
        db.close()

@app.route('/api/keywords', methods=['POST'])
@login_required
def add_keyword():
    """Add new keyword with auto-translation to 7 languages + expansion"""
    MAX_KEYWORDS_PER_USER = 20
    user_id = getattr(current_user, 'id', None)  # Get user_id early for keyword sharing
    
    data = request.get_json() or {}
    db = get_db()
    try:
        keyword_ar = (data.get('text_ar') or '').strip()

        if not keyword_ar:
            print(f"❌ Keyword add failed: empty keyword text")
            return jsonify({"error": "النص العربي للكلمة مطلوب"}), 400
        
        # Check keyword limit (5 per user)
        current_count = scoped(db.query(Keyword), Keyword, force_user_filter=True).count()
        if current_count >= MAX_KEYWORDS_PER_USER:
            print(f"❌ Keyword add failed: user {current_user.id} reached limit ({current_count}/{MAX_KEYWORDS_PER_USER})")
            return jsonify({"error": f"لقد وصلت للحد الأقصى ({MAX_KEYWORDS_PER_USER} كلمات). احذف كلمة لإضافة أخرى."}), 400
        
        # Check if keyword already exists FOR THIS USER only (force_user_filter=True)
        existing = scoped(db.query(Keyword), Keyword, force_user_filter=True).filter(Keyword.text_ar == keyword_ar).first()
        if existing:
            print(f"❌ Keyword add failed: '{keyword_ar}' already exists for user {current_user.id}")
            return jsonify({"error": "هذه الكلمة موجودة بالفعل لهذا المستخدم"}), 400
        
        print(f"✅ Adding new keyword: '{keyword_ar}' for user {current_user.id}")
        
        # Step 1: Translate keyword using Google Translate to 7 base languages (for DB storage)
        print(f"🔄 Step 1: Translating keyword to 7 base languages: {keyword_ar}")
        translations = translate_keyword(keyword_ar)
        print(f"   ✅ Translated to: en, fr, tr, ur, zh, ru, es")
        
        # Create keyword with all translations
        keyword = Keyword(
            text_ar=keyword_ar,
            text_en=translations.get('en'),
            text_fr=translations.get('fr'),
            text_tr=translations.get('tr'),
            text_ur=translations.get('ur'),
            text_zh=translations.get('zh'),
            text_ru=translations.get('ru'),
            text_es=translations.get('es'),
            enabled=True,
            user_id=getattr(current_user, 'id', None)
        )
        
        db.add(keyword)
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            error_str = str(e).lower()
            print(f"❌ Database error adding keyword: {e}")
            # Handle existing UNIQUE constraint (SQLite or PostgreSQL)
            if 'unique' in error_str or 'duplicate' in error_str:
                return jsonify({"error": "هذه الكلمة موجودة مسبقاً"}), 400
            raise
        
        print(f"   ✅ Keyword saved to database (ID: {keyword.id})")
        
        # Step 1.5: Check if other users have this keyword and copy their articles
        copied_count = copy_articles_from_shared_keyword(db, keyword_ar, user_id)
        if copied_count > 0:
            print(f"   📋 Copied {copied_count} articles from other users with same keyword")
        
        # Step 2: Expand keyword to 32 languages and SAVE TO DATABASE
        print(f"🔄 Step 2: Expanding keyword to 32 languages for global search...")
        expansion = expand_keyword(keyword_ar, keyword_obj=keyword, db=db)
        
        if expansion['status'] == 'success':
            print(f"   ✅ Expanded successfully to {len(expansion['translations'])} languages (saved to DB)")
            print(f"   📋 Coverage: en, fr, es, de, ru, zh-cn, ja, hi, id, pt, tr, ko, it, nl, vi, th, ms, fa, ur, +more")
        elif expansion['status'] == 'partial':
            print(f"   ⚠️  Partially expanded to {len(expansion['translations'])} languages (saved to DB)")
        else:
            print(f"   ❌ Expansion failed")
        
        # Step 3: Ensure global scheduler is running and trigger immediate run
        if user_id:
            status = global_scheduler.get_status()
            if not status.get('running'):
                print(f"🚀 Starting global scheduler (first keyword added)...")
                global_scheduler.start()
            else:
                print(f"🔄 Triggering immediate monitoring run for new keyword...")
                global_scheduler.trigger_now()
            print(f"   ✅ Global monitoring active")
        
        return jsonify({
            "success": True,
            "id": keyword.id,
            "translations": translations,
            "expansion": expansion  # Include expansion in response
        })
    finally:
        db.close()

@app.route('/api/keywords/<int:keyword_id>', methods=['DELETE'])
@login_required
def delete_keyword(keyword_id):
    """Delete keyword - stops monitoring if no keywords remain"""
    db = get_db()
    try:
        query = scoped(db.query(Keyword), Keyword, force_user_filter=True)
        keyword = query.filter(Keyword.id == keyword_id).first()
        
        if not keyword:
            return jsonify({"error": "Keyword not found"}), 404
        
        db.delete(keyword)
        db.commit()
        
        # Check if there are any remaining enabled keywords for this user
        user_id = getattr(current_user, 'id', None)
        remaining_keywords = scoped(db.query(Keyword), Keyword, force_user_filter=True).filter(Keyword.enabled == True).count()
        
        # Note: Global scheduler keeps running for other users
        # It will skip users with no keywords automatically
        if remaining_keywords == 0 and user_id:
            print(f"ℹ️ User {user_id} has no keywords remaining - will be skipped in next global run")
        
        return jsonify({"success": True, "remaining_keywords": remaining_keywords})
    finally:
        db.close()

@app.route('/api/keywords/<int:keyword_id>/toggle', methods=['POST'])
@login_required
def toggle_keyword(keyword_id):
    """Toggle keyword enabled status"""
    db = get_db()
    try:
        query = scoped(db.query(Keyword), Keyword, force_user_filter=True)
        keyword = query.filter(Keyword.id == keyword_id).first()
        
        if not keyword:
            return jsonify({"error": "Keyword not found"}), 404
        
        keyword.enabled = not keyword.enabled
        db.commit()
        
        return jsonify({"success": True, "enabled": keyword.enabled})
    finally:
        db.close()

# ==================== Monitoring Jobs (Background Execution) ====================

from job_executor import job_executor

@app.route('/api/monitor/job/start', methods=['POST'])
@login_required
def start_monitor_job():
    """
    Start a new monitoring job (background execution).
    
    Returns immediately with job_id. Use /api/monitor/job/status to poll.
    
    Features:
    - Non-blocking (returns in <100ms)
    - Idempotent (returns existing job if one is running)
    - Rate limited (max 10 jobs/hour per user)
    - Per-user isolation
    
    Response:
        {
            "success": true,
            "job_id": 123,
            "status": "QUEUED",
            "message": "Monitoring job started",
            "existing": false
        }
    """
    user_id = current_user.id
    result = job_executor.start_monitoring_job(user_id)
    
    if result.get('success'):
        return jsonify(result)
    else:
        return jsonify(result), 429 if 'Rate limit' in result.get('error', '') else 503


@app.route('/api/monitor/job/status', methods=['GET'])
@login_required
def get_monitor_job_status():
    """
    Get status of active monitoring job for current user.
    
    Response:
        {
            "id": 123,
            "status": "RUNNING",
            "progress": 45,
            "progress_message": "Fetching RSS feeds...",
            "total_fetched": 500,
            "total_matched": 25,
            "total_saved": 0,
            "started_at": "2026-01-12T12:00:00",
            ...
        }
    """
    user_id = current_user.id
    
    # First check for active job
    active_job = job_executor.get_active_job(user_id)
    if active_job:
        return jsonify(active_job)
    
    # No active job - return most recent
    recent_jobs = job_executor.get_user_jobs(user_id, limit=1)
    if recent_jobs:
        return jsonify(recent_jobs[0])
    
    return jsonify({"status": "NONE", "message": "No jobs found"})


@app.route('/api/monitor/job/<int:job_id>', methods=['GET'])
@login_required
def get_monitor_job_by_id(job_id):
    """Get status of a specific job by ID"""
    user_id = current_user.id
    job = job_executor.get_job_status(job_id, user_id)
    
    if not job:
        return jsonify({"error": "Job not found"}), 404
    
    return jsonify(job)


@app.route('/api/monitor/job/<int:job_id>/cancel', methods=['POST'])
@login_required
def cancel_monitor_job(job_id):
    """Cancel a running monitoring job"""
    user_id = current_user.id
    result = job_executor.cancel_job(job_id, user_id)
    
    if result.get('success'):
        return jsonify(result)
    else:
        return jsonify(result), 400


@app.route('/api/monitor/jobs', methods=['GET'])
@login_required
def get_monitor_jobs_history():
    """Get recent monitoring jobs for current user"""
    user_id = current_user.id
    limit = request.args.get('limit', 10, type=int)
    limit = min(limit, 50)  # Cap at 50
    
    jobs = job_executor.get_user_jobs(user_id, limit=limit)
    return jsonify({"jobs": jobs})


# ==================== Monitoring Scheduler (Global) ====================

from global_scheduler import global_scheduler

@app.route('/api/monitor/status', methods=['GET'])
@login_required
def get_monitor_status():
    """Get monitoring status for current user (backed by global scheduler).
    
    Also acts as a WATCHDOG: if the scheduler thread crashed, this will
    detect it and restart it automatically (called every 30s by Dashboard).
    """
    # Watchdog: auto-restart dead scheduler thread
    global_scheduler.ensure_running()
    
    user_id = current_user.id
    return jsonify(global_scheduler.get_user_status(user_id))

@app.route('/api/monitor/start', methods=['POST'])
@admin_required
def start_monitoring_scheduler():
    """Start the global monitoring scheduler (admin action, benefits all users)"""
    result = global_scheduler.start()
    return jsonify(result)

@app.route('/api/monitor/stop', methods=['POST'])
@admin_required
def stop_monitoring_scheduler():
    """Stop the global monitoring scheduler (admin action)"""
    result = global_scheduler.stop()
    return jsonify(result)

# ==================== Monitoring (Manual) ====================

@app.route('/api/monitor/run', methods=['POST'])
@login_required
def run_monitoring():
    """
    Main monitoring function - fetches news and analyzes with Gemini
    This runs on-demand when user clicks the button
    
    SECURITY: Only uses CURRENT USER's keywords for monitoring.
    Results are saved with current user's ID for isolation.
    """
    db = get_db()
    user_id = current_user.id
    
    try:
        # Get ALL enabled sources (shared catalog)
        sources = db.query(Source).filter(Source.enabled == True).all()
        
        # SECURITY FIX: Get ONLY current user's enabled keywords
        keywords = db.query(Keyword).filter(
            Keyword.enabled == True,
            Keyword.user_id == user_id
        ).all()
        
        if not sources:
            return jsonify({"error": "No enabled sources"}), 400
        
        if not keywords:
            return jsonify({"error": "No enabled keywords"}), 400
        
        # Log sources breakdown by country (for transparency)
        print("\n📊 Enabled Sources Breakdown:")
        country_counts = {}
        for s in sources:
            country = s.country_name
            country_counts[country] = country_counts.get(country, 0) + 1
        
        for country, count in sorted(country_counts.items(), key=lambda x: -x[1])[:10]:
            print(f"   • {country}: {count} sources")
        
        if len(country_counts) > 10:
            print(f"   ... and {len(country_counts) - 10} more countries")
        
        print(f"\n✅ Total: {len(sources)} sources from {len(country_counts)} countries")
        print(f"✅ Keywords: {len(keywords)}")
        print(f"🎯 System will fetch from ALL countries equally (no bias!)\n")
        
        # Convert to dicts for processing
        sources_list = [{
            'id': s.id,
            'country_name': s.country_name,
            'name': s.name,
            'url': s.url,
            'enabled': s.enabled
        } for s in sources]
        
        keywords_list = [{
            'id': k.id,
            'text_ar': k.text_ar,
            'text_en': k.text_en,
            'text_fr': k.text_fr,
            'text_tr': k.text_tr,
            'text_ur': k.text_ur,
            'text_zh': k.text_zh,
            'text_ru': k.text_ru,
            'text_es': k.text_es,
            'enabled': k.enabled
        } for k in keywords]
        
        # Step 1: Load CACHED keyword expansions (NEVER translates during monitoring!)
        print("🔄 Loading cached keyword expansions...")
        keyword_expansions = load_expansions_from_keywords(keywords)
        
        if not keyword_expansions:
            print("❌ No cached expansions found!")
            print("   → Please add keywords via frontend first")
            return jsonify({"error": "No cached keyword expansions. Please add keywords first."}), 400
        
        print(f"✅ Loaded {len(keyword_expansions)} cached expansions (no translation needed)")
        print(f"   ⚡ Instant load - expansions were pre-computed when keywords were added\n")
        
        # Step 2: Fetch all feeds CONCURRENTLY and match with keywords (NEW OPTIMIZED!)
        monitoring_result = run_optimized_monitoring(
            sources_list,
            keyword_expansions,
            max_concurrent=50,  # Fetch 50 sources at once!
            max_per_source=30
        )
        
        if not monitoring_result['success'] or not monitoring_result.get('matches'):
            print("⚠️ No matches found")
            return jsonify({    
                "success": True,
                "total_fetched": monitoring_result.get('total_fetched', 0),
                "total_matches": 0,
                "articles": []
            })
        
        matches = monitoring_result['matches']
        
        # Step 3: Save matched articles with multilingual translations
        print(f"\n💾 Saving matched articles with Arabic translations...")
        print("ℹ️ Using Google Translate (FREE) with caching")
        print("ℹ️ Sentiment analysis: Neutral (default)\n")
        
        # Use new save function with balancing (no hard-coded limit!) and per-user ownership
        from flask_login import current_user as _cu
        owner_id = getattr(_cu, 'id', None)
        saved_ids, save_stats = save_matched_articles_sync(
            db,
            matches,
            apply_limit=True,
            user_id=owner_id,
        )
        
        # Retrieve saved articles for response
        processed_articles = []
        if saved_ids:
            saved_articles = db.query(Article).filter(Article.id.in_(saved_ids)).all()
            for a in saved_articles:
                processed_articles.append({
                    'id': a.id,
                    'country': a.country,
                    'source_name': a.source_name,
                    'keyword': a.keyword_original,
                    'title_ar': a.title_ar,
                    'summary_ar': a.summary_ar,
                    'sentiment': a.sentiment_label,
                    'language': a.original_language,
                    'url': a.url,
                    'image_url': a.image_url,
                    'published_at': a.published_at.isoformat() if a.published_at else None
                })
        
        print(f"\n{'='*50}")
        print(f"✅ Monitoring complete!")
        print(f"📊 Saved {len(processed_articles)} new articles")
        print(f"{'='*50}\n")
        
        # Build comprehensive response with all stats
        return jsonify({
            "success": True,
            
            # Fetch statistics
            "total_sources": len(sources_list),
            "successful_sources": len([r for r in monitoring_result.get('fetch_results', []) if r.get('status') == 'success']),
            "total_fetched": monitoring_result.get('total_fetched', 0),
            
            # Match statistics
            "total_matches": len(matches),
            "per_keyword_stats": monitoring_result.get('keyword_stats', {}),
            "acceptance_rate": monitoring_result.get('acceptance_rate', 0),
            
            # Save statistics
            "total_saved": save_stats.get('total_saved', 0),
            "duplicates_skipped": save_stats.get('duplicates_skipped', 0),
            "save_limit_applied": save_stats.get('limit_applied', False),
            "save_limit": save_stats.get('save_limit'),
            "balancing_stats": save_stats.get('balancing_stats'),
            
            # Feed health
            "feed_health": monitoring_result.get('feed_health'),
            
            # Config
            "config": monitoring_result.get('config'),
            
            # Articles
            "articles": processed_articles
        })
    
    except Exception as e:
        print(f"❌ Error in monitoring: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()

# ==================== Articles ====================

@app.route('/api/articles', methods=['GET'])
@login_required
def get_articles():
    """Get current user's articles with optional filters and pagination.
    
    SECURITY: Uses force_user_filter=True to ensure each user 
    only sees their own articles, even admins.
    
    Pagination params:
      - page: 1-indexed page number (default 1)
      - per_page: articles per page (default 30, max 100)
    
    Returns: { articles: [...], total, page, per_page, total_pages }
    """
    db = get_db()
    try:
        # Get query parameters
        country = request.args.get('country')
        keyword = request.args.get('keyword')
        sentiment = request.args.get('sentiment')
        search = request.args.get('search')
        
        # Pagination params (per_page=0 means return ALL — used by exports)
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 30, type=int)
        page = max(1, page)
        return_all = (per_page == 0)
        if not return_all:
            per_page = max(1, min(per_page, 100))
        
        # SECURITY FIX: Force user filter to ensure isolation (even for admin)
        query = scoped(db.query(Article), Article, force_user_filter=True).order_by(Article.created_at.desc())
        
        # Apply filters
        if country:
            query = query.filter(Article.country == country)
        if keyword:
            # Use keyword_original (NEW field) instead of keyword (deprecated)
            query = query.filter(Article.keyword_original == keyword)
        if sentiment:
            # Use sentiment_label (NEW field) instead of sentiment (deprecated)
            query = query.filter(Article.sentiment_label == sentiment)
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                (Article.title_ar.like(search_term)) |
                (Article.summary_ar.like(search_term))
            )
        
        # Get total count (lightweight — no data loaded)
        total = query.count()
        
        if return_all:
            # Export mode: return everything, no pagination
            per_page = total
            total_pages = 1
            page = 1
            articles = query.all()
        else:
            total_pages = max(1, (total + per_page - 1) // per_page)
            # Clamp page to valid range
            if page > total_pages:
                page = total_pages
            # Paginate: only fetch the rows we need
            offset = (page - 1) * per_page
            articles = query.offset(offset).limit(per_page).all()
        
        result = []
        for a in articles:
            # Parse keywords_translations JSON to get match context
            match_context = None
            keyword_original = a.keyword_original or a.keyword
            
            if a.keywords_translations:
                try:
                    keywords_data = json.loads(a.keywords_translations)
                    match_contexts = keywords_data.get('match_contexts', [])
                    # Get context for primary keyword
                    if match_contexts:
                        match_context = match_contexts[0]  # Use first context (primary keyword)
                except:
                    pass  # Ignore JSON parse errors
            
            article_data = {
                'id': a.id,
                'country': a.country,
                'source_name': a.source_name,
                'keyword': keyword_original,
                'keyword_original': keyword_original,
                # Original fields
                'title': a.title_original,
                'summary': a.summary_original,
                'language_detected': a.original_language or 'unknown',
                'image_url': a.image_url,
                # Arabic translations
                'title_ar': a.title_ar,
                'summary_ar': a.summary_ar,
                'translation_status': 'success' if a.title_ar and a.summary_ar else 'partial' if a.title_ar or a.summary_ar else 'none',
                # Match context (NEW - for showing why article matched)
                'match_context': match_context,
                # Other fields
                'sentiment': a.sentiment_label or a.sentiment,
                'url': a.url,
                'published_at': a.published_at.isoformat() if a.published_at else None,
                'created_at': a.created_at.isoformat() if a.created_at else None
            }
            result.append(article_data)
        
        return jsonify({
            'articles': result,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages,
        })
    finally:
        db.close()

@app.route('/api/sources/countries', methods=['GET'])
def get_sources_countries():
    """Get countries from Country table with source count (for Top Headlines)"""
    db = get_db()
    try:
        from sqlalchemy import func
        
        # Single query with GROUP BY - fixes N+1 query issue
        source_counts = db.query(
            Source.country_name,
            func.count(Source.id).label('count')
        ).filter(
            Source.enabled == True
        ).group_by(Source.country_name).all()
        
        # Convert to dict for fast lookup
        count_map = {row[0]: row[1] for row in source_counts}
        
        # Get enabled countries and join with source counts
        countries = db.query(Country).filter(Country.enabled == True).all()
        
        result = []
        for country in countries:
            source_count = count_map.get(country.name_ar, 0)
            if source_count > 0:
                result.append({
                    'name': country.name_ar,
                    'count': source_count
                })
        
        # Sort by source count descending
        result.sort(key=lambda x: x['count'], reverse=True)
        
        # Log for debugging
        if result:
            print(f"📰 Found {len(result)} countries with sources:")
            for c in result[:5]:  # Only log top 5 to reduce noise
                print(f"   • {c['name']}: {c['count']} sources")
            if len(result) > 5:
                print(f"   ... and {len(result) - 5} more countries")
        
        return jsonify({'countries': result})
    finally:
        db.close()


@app.route('/api/articles/countries', methods=['GET'])
@login_required
def get_articles_countries():
    """Get distinct countries from current user's articles.
    
    SECURITY: Filters by user_id to ensure isolation.
    """
    db = get_db()
    try:
        from sqlalchemy import distinct, func
        
        # SECURITY FIX: Filter by current user's articles only
        countries_query = db.query(
            Article.country,
            func.count(Article.id).label('article_count')
        ).filter(Article.user_id == current_user.id).group_by(Article.country).order_by(func.count(Article.id).desc())
        
        result = []
        for country, count in countries_query:
            if country:  # Skip None/empty
                result.append({
                    'name_ar': country,
                    'article_count': count
                })
        
        # Log for debugging
        if result:
            print(f"📊 Found {len(result)} countries with articles:")
            for c in result:
                print(f"   • {c['name_ar']}: {c['article_count']} articles")
        
        return jsonify(result)
    finally:
        db.close()

@app.route('/api/health/newsdata', methods=['GET'])
def health_check_newsdata():
    """Test NewsData.io API key — no auth required."""
    from newsdata_client import _read_env_key
    key = _read_env_key('NEWSDATA_API_KEY')
    if not key:
        return jsonify({"ok": False, "error": "Key not found in .env"}), 500
    try:
        import requests as _r
        resp = _r.get('https://newsdata.io/api/1/latest', params={'apikey': key, 'q': 'test', 'size': 1}, timeout=10)
        data = resp.json()
        return jsonify({
            "ok": data.get('status') == 'success',
            "key_prefix": key[:12] + '...',
            "api_status": data.get('status'),
            "api_message": data.get('message', ''),
            "total": data.get('totalResults', 0),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)[:200]}), 500

@app.route('/api/health/translation', methods=['GET'])
def health_check_translation():
    """Test Google Translate service (no API key needed)"""
    try:
        from deep_translator import GoogleTranslator
        
        # Quick test: translate "hello" to Arabic
        result = GoogleTranslator(source='en', target='ar').translate("hello")
        
        if result:
            return jsonify({
                "ok": True,
                "service": "Google Translate",
                "test": f"hello → {result}",
                "status": "FREE - No API key required"
            })
        else:
            return jsonify({
                "ok": False,
                "error": "No response from Google Translate"
            }), 500
            
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@app.route('/api/translate-text', methods=['POST'])
@login_required
def translate_text_endpoint():
    """Translate title and/or summary to Arabic using Google Translate (free)."""
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    summary = (data.get('summary') or '').strip()

    if not title and not summary:
        return jsonify({'error': 'No text provided'}), 400

    try:
        from deep_translator import GoogleTranslator

        def _is_arabic(text):
            if not text:
                return True
            arabic_count = sum(1 for c in text if '\u0600' <= c <= '\u06FF' or '\u0750' <= c <= '\u077F')
            return arabic_count / max(len(text), 1) > 0.3

        title_ar = title
        summary_ar = summary

        if title and not _is_arabic(title):
            title_ar = GoogleTranslator(source='auto', target='ar').translate(title) or title
        if summary and not _is_arabic(summary):
            summary_ar = GoogleTranslator(source='auto', target='ar').translate(summary[:1000]) or summary

        return jsonify({
            'title_ar': title_ar,
            'summary_ar': summary_ar,
            'translated': (title_ar != title) or (summary_ar != summary),
        })
    except Exception as e:
        print(f"[Translate] error: {e}")
        return jsonify({'error': str(e)[:200]}), 500

@app.route('/api/articles/stats', methods=['GET'])
@login_required
def get_article_stats():
    """Get article statistics for current user only.
    
    SECURITY: Filters by user_id to ensure isolation.
    """
    db = get_db()
    try:
        from sqlalchemy import or_
        user_id = current_user.id
        
        # SECURITY FIX: Filter all counts by user_id
        base_query = db.query(Article).filter(Article.user_id == user_id)
        
        total = base_query.count()
        
        positive = base_query.filter(
            or_(
                Article.sentiment_label == 'إيجابي',
                Article.sentiment == 'إيجابي'
            )
        ).count()
        
        negative = base_query.filter(
            or_(
                Article.sentiment_label == 'سلبي',
                Article.sentiment == 'سلبي'
            )
        ).count()
        
        neutral = base_query.filter(
            or_(
                Article.sentiment_label == 'محايد',
                Article.sentiment == 'محايد'
            )
        ).count()
        
        # Count unique countries with results
        unique_countries = db.query(Article.country).filter(
            Article.user_id == user_id
        ).distinct().count()
        
        return jsonify({
            'total': total,
            'positive': positive,
            'negative': negative,
            'neutral': neutral,
            'uniqueCountries': unique_countries
        })
    finally:
        db.close()

@app.route('/api/articles/clear', methods=['POST'])
@login_required
def clear_articles():
    """Clear current user's articles only.
    
    SECURITY: Only deletes articles belonging to current user.
    """
    db = get_db()
    try:
        # SECURITY FIX: Only delete current user's articles
        deleted = db.query(Article).filter(Article.user_id == current_user.id).delete()
        db.commit()
        
        return jsonify({"success": True, "deleted": deleted})
    finally:
        db.close()

@app.route('/api/articles/export-and-reset', methods=['POST'])
@login_required
def export_and_reset():
    """
    Export current user's articles to Excel, then delete user's articles and keywords.
    Atomic operation: all or nothing.
    
    SECURITY: Only exports/deletes data belonging to current user.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from datetime import datetime as dt
    import os
    
    from io import BytesIO
    
    db = get_db()
    user_id = current_user.id
    
    try:
        # Step 1: Fetch ONLY current user's articles
        print(f"📊 Fetching articles for user {user_id}...")
        articles = db.query(Article).filter(Article.user_id == user_id).order_by(Article.created_at.desc()).all()
        article_count = len(articles)
        
        if article_count == 0:
            return jsonify({"error": "No articles to export"}), 400
        
        print(f"   ✅ Found {article_count} articles")
        
        # Step 2: Create Excel file in memory (no filesystem — Render compatible)
        timestamp = dt.now().strftime('%Y-%m-%d_%H-%M-%S')
        filename = f"export_{timestamp}.xlsx"
        
        print(f"📝 Creating Excel file: {filename}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "articles"
        
        # Headers - matching spec exactly
        headers = [
            'id', 'title_original', 'title_ar', 'source', 'country', 'url', 
            'published_at_utc', 'original_language', 'arabic_text', 
            'keyword_original', 'keywords_translations', 'sentiment_label', 
            'sentiment_score', 'fetched_at_utc'
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write article data
        for row_idx, article in enumerate(articles, 2):
            ws.cell(row=row_idx, column=1, value=article.id)
            ws.cell(row=row_idx, column=2, value=article.title_original or '')
            ws.cell(row=row_idx, column=3, value=article.title_ar or '')
            ws.cell(row=row_idx, column=4, value=article.source_name)
            ws.cell(row=row_idx, column=5, value=article.country)
            ws.cell(row=row_idx, column=6, value=article.url)
            ws.cell(row=row_idx, column=7, value=article.published_at.isoformat() if article.published_at else '')
            ws.cell(row=row_idx, column=8, value=article.original_language or article.language or '')
            ws.cell(row=row_idx, column=9, value=article.arabic_text or (article.title_ar + ' ' + (article.summary_ar or '')))
            ws.cell(row=row_idx, column=10, value=article.keyword_original or article.keyword or '')
            ws.cell(row=row_idx, column=11, value=article.keywords_translations or '')
            ws.cell(row=row_idx, column=12, value=article.sentiment_label or article.sentiment or '')
            ws.cell(row=row_idx, column=13, value=article.sentiment_score or '')
            ws.cell(row=row_idx, column=14, value=article.fetched_at.isoformat() if article.fetched_at else article.created_at.isoformat())
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook to memory buffer
        buf = BytesIO()
        wb.save(buf)
        file_data = buf.getvalue()
        buf.close()
        print(f"   ✅ Excel file created in memory: {len(file_data)} bytes")
        
        # Step 3: Store in database as ExportRecord (uses existing download route)
        rec = ExportRecord(
            user_id=user_id,
            filters_json=json.dumps({"type": "export_and_reset"}, ensure_ascii=False),
            article_count=article_count,
            filename=filename,
            file_data=file_data,
            file_size=len(file_data),
            source_type='export_reset',
        )
        db.add(rec)
        db.flush()  # Get rec.id before commit
        export_id = rec.id
        print(f"   ✅ Export record created: id={export_id}")
        
        # Step 4: Delete ONLY current user's data (atomic transaction)
        print(f"🗑️ Starting atomic delete transaction for user {user_id}...")
        
        # SECURITY FIX: Delete only current user's data
        deleted_articles = db.query(Article).filter(Article.user_id == user_id).delete()
        deleted_keywords = db.query(Keyword).filter(Keyword.user_id == user_id).delete()
        db.commit()
        
        print(f"   ✅ Deleted {deleted_articles} articles and {deleted_keywords} keywords for user {user_id}")
        
        # Step 5: Clear in-memory caches
        from translation_service import clear_all_caches
        clear_all_caches()
        
        # Step 6: Return success with correct download URL
        return jsonify({
            "success": True,
            "filename": filename,
            "article_count": article_count,
            "download_url": f"/api/exports/{export_id}/download"
        })
        
    except Exception as e:
        print(f"❌ Export & Reset failed: {str(e)}")
        db.rollback()
        return jsonify({"error": f"Export failed: {str(e)}"}), 500
        
    finally:
        db.close()

# ==================== Diagnostics ====================

@app.route('/api/feeds/diagnose', methods=['GET'])
@admin_required
def diagnose_feeds():
    """
    Diagnose all feeds - check their status and errors
    Returns detailed status for each feed
    """
    db = get_db()
    try:
        # Get limit from query params (0 = all)
        limit = int(request.args.get('limit', '0'))
        
        # Get all sources or limited set
        query = db.query(Source).order_by(Source.country_name, Source.name)
        if limit > 0:
            query = query.limit(limit)
        sources = query.all()
        
        results = []
        count = 0
        
        print(f"\n{'='*50}")
        print(f"🔍 Diagnosing {len(sources)} feeds...")
        print(f"{'='*50}\n")
        
        for source in sources:
            print(f"Testing: {source.name} ({source.country_name})")
            
            result = fetch_feed(source.url)
            
            results.append({
                "country": source.country_name,
                "source": source.name,
                "url": source.url,
                "enabled": source.enabled,
                "status": result["status"],
                "http_status": result["http_status"],
                "entries_count": len(result["entries"]),
                "error": result["error"]
            })
            
            count += 1
            
            # Small delay between tests
            import time
            time.sleep(0.3)
        
        print(f"\n{'='*50}")
        print(f"✅ Diagnosis complete!")
        print(f"{'='*50}\n")
        
        return jsonify({
            "success": True,
            "total_tested": count,
            "feeds": results
        })
    finally:
        db.close()

@app.route('/api/feeds/selftest', methods=['GET'])
@admin_required
def selftest_feeds():
    """
    Quick self-test - test first N enabled feeds
    Returns summary of working vs failing feeds
    """
    db = get_db()
    try:
        # Get limit from query params (default 5)
        limit = int(request.args.get('limit', '5'))
        
        # Get first N enabled sources
        sources = db.query(Source).filter(Source.enabled == True).limit(limit).all()
        
        if not sources:
            return jsonify({
                "success": False,
                "error": "No enabled sources found"
            }), 400
        
        results = []
        working = 0
        failing = 0
        
        print(f"\n{'='*50}")
        print(f"🧪 Self-test: Testing {len(sources)} feeds...")
        print(f"{'='*50}\n")
        
        for source in sources:
            print(f"Testing: {source.name}")
            
            result = fetch_feed(source.url)
            
            is_ok = result["status"] == "ok" and len(result["entries"]) > 0
            
            if is_ok:
                working += 1
                print(f"   ✅ OK - {len(result['entries'])} entries")
            else:
                failing += 1
                print(f"   ❌ {result['status']} - {result['error']}")
            
            results.append({
                "source": source.name,
                "country": source.country_name,
                "status": result["status"],
                "entries": len(result["entries"]),
                "ok": is_ok
            })
            
            import time
            time.sleep(0.3)
        
        print(f"\n{'='*50}")
        print(f"✅ Self-test complete!")
        print(f"   Working: {working}/{len(sources)}")
        print(f"   Failing: {failing}/{len(sources)}")
        print(f"{'='*50}\n")
        
        return jsonify({
            "success": True,
            "total_tested": len(sources),
            "working": working,
            "failing": failing,
            "feeds": results
        })
    finally:
        db.close()


def calculate_relevance_score(article_data, keywords):
    """
    Calculate relevance score (0-1) for an article based on keyword presence
    
    Args:
        article_data: Dict with 'title', 'description', 'content' fields
        keywords: List of keyword variants (Arabic + English)
    
    Returns:
        Float score (0-1), higher = more relevant
    """
    title = (article_data.get('title') or '').lower()
    description = (article_data.get('description') or article_data.get('content') or '').lower()
    
    # Normalize keywords for matching
    keywords_lower = [k.lower() for k in keywords if k]
    
    score = 0.0
    matches_found = False
    
    # Check title (weight: 0.6)
    for kw in keywords_lower:
        if kw in title:
            # Title match is most important
            # Count occurrences (more = better, but capped at 3)
            occurrences = min(title.count(kw), 3)
            score += 0.6 * (occurrences / 3.0)
            matches_found = True
            break  # One match in title is enough
    
    # Check description/content (weight: 0.4)
    for kw in keywords_lower:
        if kw in description:
            # Count occurrences in description
            occurrences = min(description.count(kw), 5)
            score += 0.4 * (occurrences / 5.0)
            matches_found = True
            break
    
    # If no matches at all, score = 0
    if not matches_found:
        return 0.0
    
    # Normalize score to 0-1 range
    return min(score, 1.0)


import requests

@app.route('/api/direct-search', methods=['GET'])
@login_required
def direct_search():
    """
    Simple NewsData.io direct search:
    - Query params: q, qInTitle, timeframe, country, language, page
    - Single API call (no caching, no translation, no scoring)
    - Returns normalized results in الخلاصة-style format
    """

    API_KEY = os.getenv('NEWSDATA_API_KEY', '').strip()
    BASE_URL = 'https://newsdata.io/api/1/latest'

    # --- Basic validation ---
    keyword   = (request.args.get('q', '') or '').strip()
    next_page = (request.args.get('page', '') or '').strip()

    if not keyword and not next_page:
        return jsonify({
            "error": "الرجاء إدخال كلمة البحث",
            "results": [],
            "nextPage": None,
            "totalResults": 0
        }), 400

    if not API_KEY:
        return jsonify({
            "error": "مفتاح API غير موجود أو غير صحيح - يرجى إضافة NEWSDATA_API_KEY في ملف .env",
            "results": [],
            "nextPage": None,
            "totalResults": 0
        }), 500

    # --- Build API params (recommended simple usage) ---
    params = {"apikey": API_KEY}

    if next_page:
        # Pagination: فقط نرسل page + apikey
        params["page"] = next_page
    else:
        q_in_title = (request.args.get('qInTitle', '') or '').lower() == 'true'
        timeframe  = (request.args.get('timeframe', '') or '').strip()
        country    = (request.args.get('country', '') or '').strip()
        language   = (request.args.get('language', '') or '').strip()

        # Either q or qInTitle (كما هو موصى به)
        if q_in_title:
            params["qInTitle"] = keyword
        else:
            params["q"] = keyword

        if timeframe:
            params["timeframe"] = timeframe

        if country:
            # NewsData.io تسمح حتى 5 دول فقط
            params["country"] = ",".join(country.split(",")[:5])

        if language:
            params["language"] = language

    # --- Call NewsData.io once ---
    try:
        response = requests.get(BASE_URL, params=params, timeout=10)
    except requests.RequestException as e:
        return jsonify({
            "error": f"خطأ في الاتصال بواجهة NewsData.io: {str(e)[:200]}",
            "results": [],
            "nextPage": None,
            "totalResults": 0
        }), 200

    # --- Parse response safely ---
    try:
        data = response.json()
    except ValueError:
        return jsonify({
            "error": "خطأ في تحليل استجابة NewsData.io",
            "results": [],
            "nextPage": None,
            "totalResults": 0
        }), 200

    if response.status_code != 200 or data.get("status") != "success":
        return jsonify({
            "error": data.get("message", "فشل البحث من NewsData.io"),
            "results": [],
            "nextPage": None,
            "totalResults": 0
        }), 200

    raw_results = data.get("results", []) or []

    # --- Normalize to your الخلاصة format (بدون ترجمة أو فلترة معقدة) ---
    results = []
    for idx, article in enumerate(raw_results):
        title = (article.get("title") or "").strip()
        description = (article.get("description") or article.get("content") or "").strip()
        url = article.get("link", "")

        # تجاهل المقالات الناقصة
        if not title or not url:
            continue

        result = {
            "id": hash(url) if url else idx,
            "title_ar": title,            # بدون ترجمة – فقط تمرير العنوان كما هو
            "title_original": title,
            "summary_ar": description,    # بدون ترجمة – الوصف كما هو
            "summary_original": description,
            "url": url,
            "source_name": article.get("source_name") or article.get("source_id", "Unknown"),
            "country": article.get("country"),          # قد تكون list أو كود دولة
            "language_detected": article.get("language"),
            "published_at": article.get("pubDate"),
            "image_url": article.get("image_url"),
            "keyword_original": keyword if not next_page else "",
            "sentiment": "محايد",
            "created_at": dt.now().isoformat(),
            "is_newsdata": True
        }

        results.append(result)

    # --- Final response ---
    return jsonify({
        "results": results,
        "nextPage": data.get("nextPage"),
        "totalResults": len(results)
    }), 200


# ==================== NewsData.io Search (Basic Plan) ====================

@app.route('/api/newsdata/search', methods=['GET'])
@login_required
def newsdata_search():
    """Simple NewsData.io /latest search for Basic plan."""
    import requests as _req
    from pathlib import Path as _P

    # --- Read API key directly from .env file (bulletproof) ---
    api_key = ''
    try:
        for line in (_P(__file__).resolve().parent / '.env').read_text(encoding='utf-8').splitlines():
            if line.strip().startswith('NEWSDATA_API_KEY='):
                api_key = line.strip().split('=', 1)[1].strip()
                break
    except Exception:
        pass
    if not api_key:
        api_key = os.getenv('NEWSDATA_API_KEY', '').strip()

    if not api_key:
        return jsonify({"success": False, "error": "مفتاح API غير موجود في ملف .env", "results": []}), 500

    # --- Collect params from query string ---
    p = request.args
    api_params = {'apikey': api_key}

    # Query (mutually exclusive: q, qInTitle, qInMeta)
    if p.get('qInTitle', '').strip():
        api_params['qInTitle'] = p['qInTitle'].strip()
    elif p.get('qInMeta', '').strip():
        api_params['qInMeta'] = p['qInMeta'].strip()
    elif p.get('q', '').strip():
        api_params['q'] = p['q'].strip()

    # Basic plan filters
    if p.get('country'):
        api_params['country'] = p['country']
    if p.get('language'):
        api_params['language'] = p['language']
    if p.get('category'):
        api_params['category'] = p['category']
    if p.get('domain'):
        api_params['domain'] = p['domain']
    if p.get('excludeDomain'):
        api_params['excludedomain'] = p['excludeDomain']
    if p.get('timeframe'):
        api_params['timeframe'] = p['timeframe']
    if p.get('image') == 'true':
        api_params['image'] = 1
    if p.get('video') == 'true':
        api_params['video'] = 1
    if p.get('removeDuplicate') == 'true':
        api_params['removeduplicate'] = 1
    if p.get('page'):
        api_params['page'] = p['page']

    # Validate: need at least a query or filter
    has_query = any(k in api_params for k in ('q', 'qInTitle', 'qInMeta'))
    has_filter = any(k in api_params for k in ('country', 'language', 'category', 'domain'))
    if not has_query and not has_filter and 'page' not in api_params:
        return jsonify({"success": False, "error": "الرجاء إدخال كلمة البحث", "results": []}), 400

    # --- Call NewsData.io /latest ---
    try:
        resp = _req.get('https://newsdata.io/api/1/latest', params=api_params, timeout=15)
        data = resp.json()

        if data.get('status') != 'success':
            msg = data.get('results', {}).get('message', '') if isinstance(data.get('results'), dict) else ''
            if not msg:
                msg = data.get('message', 'خطأ من NewsData.io')
            return jsonify({"success": False, "error": msg, "results": []}), 400

        articles = data.get('results') or []
        results = []
        for art in articles:
            title = (art.get('title') or '').strip()
            url = art.get('link', '')
            if not title or not url:
                continue
            results.append({
                "title_ar": title,
                "title_original": title,
                "summary_ar": (art.get('description') or '').strip(),
                "summary_original": (art.get('description') or '').strip(),
                "url": url,
                "source_name": art.get('source_name') or art.get('source_id', ''),
                "country": art.get('country'),
                "language_detected": art.get('language'),
                "published_at": art.get('pubDate'),
                "image_url": art.get('image_url'),
                "category": art.get('category'),
                "creator": art.get('creator'),
                "source_icon": art.get('source_icon'),
                "sentiment": "محايد",
                "is_newsdata": True,
            })

        return jsonify({
            "success": True,
            "results": results,
            "totalResults": data.get('totalResults', len(results)),
            "nextPage": data.get('nextPage'),
        })

    except _req.Timeout:
        return jsonify({"success": False, "error": "انتهت مهلة الاتصال", "results": []}), 504
    except Exception as e:
        return jsonify({"success": False, "error": f"خطأ: {str(e)[:200]}", "results": []}), 500


@app.route('/api/newsdata/build-query', methods=['POST'])
@login_required
def build_query():
    """
    Build a NewsData.io query string from structured inputs.
    Returns the generated query for preview without executing search.
    """
    data = request.get_json() or {}
    
    must_include = data.get('mustInclude', [])
    any_of = data.get('anyOf', [])
    exact_phrase = data.get('exactPhrase', '')
    exclude = data.get('exclude', [])
    
    # Handle both array and comma-separated string inputs
    if isinstance(must_include, str):
        must_include = [t.strip() for t in must_include.split(',') if t.strip()]
    if isinstance(any_of, str):
        any_of = [t.strip() for t in any_of.split(',') if t.strip()]
    if isinstance(exclude, str):
        exclude = [t.strip() for t in exclude.split(',') if t.strip()]
    
    query = newsdata_client.build_query_string(
        must_include=must_include,
        any_of=any_of,
        exact_phrase=exact_phrase,
        exclude=exclude
    )
    
    return jsonify({
        "query": query,
        "length": len(query),
        "valid": len(query) <= 512  # NewsData.io query limit
    })


@app.route('/api/newsdata/sources', methods=['GET'])
@login_required
def get_newsdata_sources():
    """Get available NewsData.io sources with filters."""
    country = request.args.get('country', '').strip() or None
    language = request.args.get('language', '').strip() or None
    category = request.args.get('category', '').strip() or None
    
    result = newsdata_client.get_sources(
        country=country,
        language=language,
        category=category
    )
    
    return jsonify(result), 200 if result.get('success') else 400


@app.route('/api/headlines/top', methods=['GET'])
@login_required
def get_top_headlines():
    """
    Get top headlines from all sources in a country
    Query params: country (required), per_source (default 5), translate (default true)
    """
    from deep_translator import GoogleTranslator
    import feedparser
    from datetime import datetime as dt
    from time import mktime
    import uuid
    
    # Simple in-memory cache (60-120s TTL)
    cache_key = None
    cache_ttl = 90  # 90 seconds
    
    if not hasattr(get_top_headlines, 'cache'):
        get_top_headlines.cache = {}
    
    try:
        # Generate request ID for clearer logging
        req_id = str(uuid.uuid4())[:8]
        
        # Get params
        country_name = request.args.get('country', '').strip()
        per_source = int(request.args.get('per_source', 5))
        translate = request.args.get('translate', 'true').lower() == 'true'
        
        if not country_name:
            return jsonify({"error": "الرجاء تحديد الدولة"}), 400
        
        # Check cache
        cache_key = f"{country_name}_{per_source}_{translate}"
        if cache_key in get_top_headlines.cache:
            cached_data, cached_time = get_top_headlines.cache[cache_key]
            if (dt.now() - cached_time).total_seconds() < cache_ttl:
                print(f"[{req_id}] ✅ Cache hit for {country_name}")
                return jsonify(cached_data)
        
        print(f"[{req_id}] 🔍 Fetching top headlines for: {country_name}")
        
        # Get sources from database
        db = get_db()
        try:
            sources = db.query(Source).filter(Source.country_name == country_name).all()
            
            if not sources:
                return jsonify({"error": f"لا توجد مصادر للدولة: {country_name}"}), 404
            
            print(f"[{req_id}] Found {len(sources)} sources for {country_name}")
            
            # Fetch headlines from each source
            results = []
            for source in sources:
                source_result = {
                    'source_name': source.name,
                    'source_url': source.url,
                    'articles': [],
                    'error': None
                }
                
                try:
                    print(f"[{req_id}] 📡 {source.name}")
                    
                    # Parse RSS feed with timeout
                    import socket
                    old_timeout = socket.getdefaulttimeout()
                    socket.setdefaulttimeout(10)
                    
                    try:
                        # feedparser.parse() doesn't accept timeout parameter
                        feed = feedparser.parse(source.url)
                    except socket.timeout:
                        source_result['error'] = 'انتهت المهلة (timeout)'
                        results.append(source_result)
                        continue
                    finally:
                        socket.setdefaulttimeout(old_timeout)
                    
                    if not feed.entries:
                        source_result['error'] = 'لا توجد أخبار'
                        results.append(source_result)
                        continue
                    
                    # Sort by date (newest first) and take top N
                    entries = feed.entries[:per_source * 2]  # Fetch extra in case of issues
                    
                    # Sort by published date
                    def get_date(entry):
                        try:
                            if hasattr(entry, 'published_parsed') and entry.published_parsed:
                                return mktime(entry.published_parsed)
                            elif hasattr(entry, 'updated_parsed') and entry.updated_parsed:
                                return mktime(entry.updated_parsed)
                            else:
                                return 0
                        except:
                            return 0
                    
                    entries.sort(key=get_date, reverse=True)
                    entries = entries[:per_source]
                    
                    # Process entries
                    for entry in entries:
                        title = entry.get('title', '')
                        summary_raw = entry.get('summary', '') or entry.get('description', '')
                        url = entry.get('link', '')
                        
                        if not title or not url:
                            continue
                        
                        # Clean HTML from summary (remove ads, navigation, social links, etc.)
                        summary = clean_html_content(summary_raw) if summary_raw else ''
                        
                        # Get published date
                        pub_date = None
                        if hasattr(entry, 'published_parsed') and entry.published_parsed:
                            pub_date = dt.fromtimestamp(mktime(entry.published_parsed)).isoformat()
                        elif hasattr(entry, 'updated_parsed') and entry.updated_parsed:
                            pub_date = dt.fromtimestamp(mktime(entry.updated_parsed)).isoformat()
                        
                        # Translate if needed
                        title_ar = title
                        summary_ar = summary
                        
                        if translate:
                            try:
                                # Detect language
                                lang = 'en'
                                try:
                                    from langdetect import detect
                                    lang = detect(title)
                                except:
                                    pass
                                
                                if lang != 'ar':
                                    # Translate title
                                    if title:
                                        title_ar = GoogleTranslator(source=lang, target='ar').translate(title) or title
                                    
                                    # Translate summary (limit length)
                                    if summary:
                                        summary_ar = GoogleTranslator(source=lang, target='ar').translate(summary[:500]) or summary
                            except Exception as trans_err:
                                print(f"      Translation error: {trans_err}")
                                # Keep original if translation fails
                                pass
                        
                        # Build article object
                        article = {
                            'id': hash(url),
                            'title_ar': title_ar,
                            'title_original': title,
                            'summary_ar': summary_ar,
                            'summary_original': summary,
                            'url': url,
                            'published_at': pub_date,
                            'sentiment': 'محايد',
                            'keyword_original': '',  # No keyword for top headlines
                        }
                        
                        source_result['articles'].append(article)
                    
                    if len(source_result['articles']) > 0:
                        print(f"[{req_id}]    ✅ {len(source_result['articles'])} articles")
                    
                except Exception as source_err:
                    print(f"[{req_id}]    ❌ Error: {source_err}")
                    source_result['error'] = str(source_err)[:100]
                
                results.append(source_result)
            
            # Filter out sources with no articles
            results = [r for r in results if len(r['articles']) > 0 or r['error']]
            
            response_data = {
                'country': country_name,
                'sources': results,
                'total_sources': len(results),
                'total_articles': sum(len(r['articles']) for r in results)
            }
            
            # Update cache
            get_top_headlines.cache[cache_key] = (response_data, dt.now())
            
            # Clean old cache entries (keep only last 10)
            if len(get_top_headlines.cache) > 10:
                oldest_key = min(get_top_headlines.cache.keys(), 
                               key=lambda k: get_top_headlines.cache[k][1])
                del get_top_headlines.cache[oldest_key]
            
            print(f"[{req_id}] ✅ {country_name}: {response_data['total_articles']} articles from {response_data['total_sources']} sources")

            # Return the response data
            return jsonify(response_data)

        finally:
            db.close()

    except Exception as e:
        print(f"❌ Error in get_top_headlines: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/external/headlines', methods=['POST'])
def external_headlines():
    """External API: Get top headlines (أهم العناوين) for a given country.

    - Method: POST
    - Auth: X-API-Key header
    - Body (JSON): {"country": "اسم الدولة بالعربية", "per_source": 5, "translate": true}

    Response JSON has the same structure as /api/headlines/top:
    {
      "country": "...",
      "sources": [
        {
          "source_name": "...",
          "source_url": "...",
          "articles": [ ... ],
          "error": null | "..."
        }
      ],
      "total_sources": N,
      "total_articles": M
    }
    """

    from deep_translator import GoogleTranslator
    import feedparser
    from datetime import datetime as dt
    from time import mktime
    import socket

    # ---------- Authentication ----------
    external_key = os.getenv('EXTERNAL_API_KEY', '').strip()
    if not external_key:
        return jsonify({
            "error": "EXTERNAL_API_KEY is not configured on the server",
            "code": "config_error"
        }), 500

    # H7 FIX: Accept API key only via header (never query param — URLs leak in logs)
    provided_key = (request.headers.get('X-API-Key') or '').strip()

    # H8 FIX: Constant-time comparison to prevent timing side-channel attacks
    if not provided_key or not hmac.compare_digest(provided_key, external_key):
        return jsonify({
            "error": "Unauthorized",
            "code": "unauthorized"
        }), 401

    # ---------- Input validation ----------
    try:
        payload = request.get_json(force=True, silent=False) or {}
    except Exception:
        return jsonify({
            "error": "Invalid JSON body",
            "code": "bad_request"
        }), 400

    country_name = (payload.get('country') or '').strip()
    if not country_name:
        return jsonify({
            "error": "Field 'country' is required",
            "code": "missing_country"
        }), 400

    try:
        per_source = int(payload.get('per_source') or 5)
    except Exception:
        per_source = 5

    translate = bool(payload.get('translate', True))

    # ---------- Core logic (similar to get_top_headlines, no cache) ----------
    db = get_db()
    try:
        # First, resolve the country row by its Arabic name
        from models import Country as _Country
        country_row = db.query(_Country).filter(_Country.name_ar == country_name).first()

        if not country_row:
            return jsonify({
                "error": f"لا توجد مصادر للدولة: {country_name}",
                "code": "no_sources"
            }), 404

        # Then fetch all sources linked to this country_id
        sources = db.query(Source).filter(Source.country_id == country_row.id).all()

        if not sources:
            return jsonify({
                "error": f"لا توجد مصادر للدولة: {country_name}",
                "code": "no_sources"
            }), 404

        results = []

        for source in sources:
            source_result = {
                'source_name': source.name,
                'source_url': source.url,
                'articles': [],
                'error': None
            }

            try:
                old_timeout = socket.getdefaulttimeout()
                socket.setdefaulttimeout(10)
                try:
                    feed = feedparser.parse(source.url)
                finally:
                    socket.setdefaulttimeout(old_timeout)

                if not feed.entries:
                    source_result['error'] = 'لا توجد أخبار'
                    results.append(source_result)
                    continue

                entries = feed.entries[:per_source * 2]

                def get_date(entry):
                    try:
                        if hasattr(entry, 'published_parsed') and entry.published_parsed:
                            return mktime(entry.published_parsed)
                        if hasattr(entry, 'updated_parsed') and entry.updated_parsed:
                            return mktime(entry.updated_parsed)
                        return 0
                    except Exception:
                        return 0

                entries.sort(key=get_date, reverse=True)
                entries = entries[:per_source]

                for entry in entries:
                    title = entry.get('title', '')
                    summary_raw = entry.get('summary', '') or entry.get('description', '')
                    url = entry.get('link', '')

                    if not title or not url:
                        continue

                    summary = clean_html_content(summary_raw) if summary_raw else ''

                    # Try to extract an image URL from common RSS fields
                    image_url = None
                    try:
                        # media:content
                        media_content = getattr(entry, 'media_content', None) or entry.get('media_content')
                        if media_content and isinstance(media_content, list):
                            for m in media_content:
                                if isinstance(m, dict) and m.get('url'):
                                    image_url = m['url']
                                    break

                        # media:thumbnail
                        if not image_url:
                            media_thumb = getattr(entry, 'media_thumbnail', None) or entry.get('media_thumbnail')
                            if media_thumb and isinstance(media_thumb, list):
                                for m in media_thumb:
                                    if isinstance(m, dict) and m.get('url'):
                                        image_url = m['url']
                                        break

                        # enclosure
                        if not image_url:
                            enclosures = getattr(entry, 'enclosures', None) or entry.get('enclosures')
                            if enclosures and isinstance(enclosures, list):
                                for enc in enclosures:
                                    if isinstance(enc, dict):
                                        enc_url = enc.get('url') or enc.get('href')
                                        enc_type = enc.get('type', '')
                                    else:
                                        enc_url = getattr(enc, 'href', None) or getattr(enc, 'url', None)
                                        enc_type = getattr(enc, 'type', '')
                                    if enc_url and ('image' in (enc_type or '') or enc_url.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))):
                                        image_url = enc_url
                                        break

                        # Fallback: some feeds put image URL in a generic image field
                        if not image_url:
                            image_url = entry.get('image') or entry.get('img')

                        # Final fallback: try to parse first <img src="..."> from the raw HTML summary
                        if not image_url and summary_raw:
                            try:
                                m = re.search(r'<img[^>]+src=["\\\']([^"\\\']+)["\\\']', summary_raw, re.IGNORECASE)
                                if m:
                                    image_url = m.group(1)
                            except Exception:
                                pass
                    except Exception:
                        image_url = None

                    pub_date = None
                    if hasattr(entry, 'published_parsed') and entry.published_parsed:
                        pub_date = dt.fromtimestamp(mktime(entry.published_parsed)).isoformat()
                    elif hasattr(entry, 'updated_parsed') and entry.updated_parsed:
                        pub_date = dt.fromtimestamp(mktime(entry.updated_parsed)).isoformat()

                    title_ar = title
                    summary_ar = summary

                    if translate and translator:
                        try:
                            lang = 'en'
                            try:
                                from langdetect import detect
                                lang = detect(title)
                            except Exception:
                                pass

                            if lang != 'ar':
                                if title:
                                    title_ar = GoogleTranslator(source=lang, target='ar').translate(title) or title
                                if summary:
                                    summary_ar = GoogleTranslator(source=lang, target='ar').translate(summary[:500]) or summary
                        except Exception:
                            # Keep original text if translation fails
                            pass

                    article = {
                        'id': hash(url),
                        'title_ar': title_ar,
                        'title_original': title,
                        'summary_ar': summary_ar,
                        'summary_original': summary,
                        'url': url,
                        'image_url': image_url,
                        'published_at': pub_date,
                        'sentiment': 'محايد',
                        'keyword_original': ''
                    }

                    source_result['articles'].append(article)

                results.append(source_result)

            except Exception as source_err:
                source_result['error'] = str(source_err)[:120]
                results.append(source_result)

        # Filter out completely empty sources
        results = [r for r in results if r['articles'] or r['error']]

        response_data = {
            'country': country_name,
            'sources': results,
            'total_sources': len(results),
            'total_articles': sum(len(r['articles']) for r in results)
        }

        return jsonify(response_data), 200

    except Exception:
        return jsonify({
            "error": "Internal server error",
            "code": "server_error"
        }), 500
    finally:
        db.close()


# =============================================================================
# AI FEATURES — ملخص ذكي + تحليل المشاعر
# =============================================================================

@app.route('/api/ai/daily-brief', methods=['POST'])
@login_required
def get_daily_brief():
    """Generate or return cached AI daily brief for the user's articles.
    
    Accepts optional 'keyword' in JSON body. When provided, only articles
    matching that keyword are used and the cache key is keyword-specific.
    """
    from ai_service import generate_daily_brief

    body = request.get_json() or {}
    today = datetime.utcnow().strftime('%Y-%m-%d')
    force = body.get('force', False)
    keyword_filter = (body.get('keyword') or '').strip()

    # Cache key: "2026-02-23" for all, "2026-02-23:السعودية" for a specific keyword
    cache_key = f"{today}:{keyword_filter}" if keyword_filter else today

    db = get_db()
    try:
        # Check cache first (unless force refresh)
        if not force:
            cached = db.query(DailyBrief).filter(
                DailyBrief.user_id == current_user.id,
                DailyBrief.date_key == cache_key
            ).first()
            if cached:
                return jsonify({
                    'content': cached.content,
                    'article_count': cached.article_count,
                    'cached': True,
                    'date': today,
                    'keyword': keyword_filter or None,
                })

        # Get today's articles for this user
        start_of_day = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        q = db.query(Article).filter(
            Article.user_id == current_user.id,
            Article.fetched_at >= start_of_day
        )
        if keyword_filter:
            q = q.filter(Article.keyword_original == keyword_filter)
        articles = q.all()

        if not articles:
            no_articles_msg = (
                f'لا توجد مقالات لكلمة "{keyword_filter}" اليوم بعد.'
                if keyword_filter
                else 'لا توجد مقالات مرصودة اليوم بعد. ستظهر الملخص عند وصول أخبار جديدة.'
            )
            return jsonify({
                'content': no_articles_msg,
                'article_count': 0,
                'cached': False,
                'date': today,
                'keyword': keyword_filter or None,
            })

        # Build article dicts for AI
        article_dicts = [{
            'title_ar': a.title_ar or a.title_original or '',
            'sentiment': a.sentiment_label or a.sentiment or '',
            'source_name': a.source_name or '',
            'keyword_original': a.keyword_original or '',
            'country': a.country or '',
        } for a in articles]

        content = generate_daily_brief(article_dicts)

        # Cache the result (keyword-specific cache key)
        existing = db.query(DailyBrief).filter(
            DailyBrief.user_id == current_user.id,
            DailyBrief.date_key == cache_key
        ).first()
        if existing:
            existing.content = content
            existing.article_count = len(articles)
            existing.created_at = datetime.utcnow()
        else:
            db.add(DailyBrief(
                user_id=current_user.id,
                date_key=cache_key,
                content=content,
                article_count=len(articles),
            ))
        db.commit()

        return jsonify({
            'content': content,
            'article_count': len(articles),
            'cached': False,
            'date': today,
            'keyword': keyword_filter or None,
        })
    except Exception as e:
        print(f"[AI] ❌ Daily brief error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/api/ai/country-brief', methods=['POST'])
@login_required
def country_brief():
    """Generate AI summary for a specific country's recent news (all users)."""
    from ai_service import generate_daily_brief

    body = request.get_json() or {}
    country = (body.get('country') or '').strip()
    if not country:
        return jsonify({'error': 'country is required'}), 400

    db = get_db()
    try:
        articles = db.query(Article).filter(
            Article.country == country
        ).order_by(Article.created_at.desc()).limit(30).all()

        if not articles:
            return jsonify({
                'content': f'لا توجد أخبار مرصودة من {country} حالياً.',
                'article_count': 0,
            })

        article_dicts = [{
            'title_ar': a.title_ar or a.title_original or '',
            'sentiment': a.sentiment_label or a.sentiment or '',
            'source_name': a.source_name or '',
            'keyword_original': a.keyword_original or '',
            'country': a.country or '',
        } for a in articles]

        content = generate_daily_brief(article_dicts)
        return jsonify({
            'content': content,
            'article_count': len(articles),
        })
    except Exception as e:
        print(f"[AI] Country brief error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/api/ai/explain-sentiment', methods=['POST'])
@login_required
def explain_sentiment_endpoint():
    """AI explains why an article has a certain sentiment — on-demand."""
    from ai_service import explain_sentiment

    data = request.get_json() or {}
    title = data.get('title', '')
    summary = data.get('summary', '')
    sentiment = data.get('sentiment', '')

    if not title:
        return jsonify({'error': 'Title is required'}), 400

    try:
        explanation = explain_sentiment(
            title=title,
            summary=summary,
            sentiment=sentiment,
            source_name=data.get('source_name', ''),
            country=data.get('country', ''),
            keyword=data.get('keyword', ''),
        )
        return jsonify({'explanation': explanation})
    except Exception as e:
        print(f"[AI] ❌ Sentiment explanation error: {e}")
        return jsonify({'error': str(e)}), 500


# =============================================================================
# BOOKMARKS — المفضلة (survives monthly reset)
# =============================================================================

@app.route('/api/bookmarks', methods=['GET'])
@login_required
def get_bookmarks():
    """Get all bookmarked articles for the current user."""
    db = get_db()
    try:
        bookmarks = db.query(BookmarkedArticle).filter(
            BookmarkedArticle.user_id == current_user.id
        ).order_by(BookmarkedArticle.bookmarked_at.desc()).all()

        return jsonify([{
            'id': b.id,
            'original_article_id': b.original_article_id,
            'title_ar': b.title_ar,
            'title_original': b.title_original,
            'summary_ar': b.summary_ar,
            'url': b.original_url,
            'image_url': b.image_url,
            'source_name': b.source_name,
            'country': b.country,
            'keyword_original': b.keyword_original,
            'sentiment': b.sentiment,
            'published_at': b.published_at.isoformat() if b.published_at else None,
            'bookmarked_at': b.bookmarked_at.isoformat() if b.bookmarked_at else None,
            'note': b.note,
        } for b in bookmarks])
    finally:
        db.close()


@app.route('/api/bookmarks', methods=['POST'])
@login_required
def create_bookmark():
    """Bookmark an article — saves a snapshot that survives monthly reset."""
    data = request.get_json() or {}
    url = data.get('url', '').strip()
    if not url:
        return jsonify({'error': 'URL is required'}), 400

    db = get_db()
    try:
        existing = db.query(BookmarkedArticle).filter(
            BookmarkedArticle.user_id == current_user.id,
            BookmarkedArticle.original_url == url
        ).first()
        if existing:
            return jsonify({'error': 'already_bookmarked', 'id': existing.id}), 409

        pub_at = None
        if data.get('published_at'):
            try:
                pub_at = datetime.fromisoformat(str(data['published_at']).replace('Z', '+00:00'))
            except Exception:
                pass

        bookmark = BookmarkedArticle(
            user_id=current_user.id,
            original_article_id=data.get('article_id'),
            title_ar=data.get('title_ar', ''),
            title_original=data.get('title_original', ''),
            summary_ar=data.get('summary_ar', ''),
            original_url=url,
            image_url=data.get('image_url'),
            source_name=data.get('source_name', ''),
            country=data.get('country', ''),
            keyword_original=data.get('keyword_original', ''),
            sentiment=data.get('sentiment', ''),
            published_at=pub_at,
        )
        db.add(bookmark)
        db.commit()
        return jsonify({'success': True, 'id': bookmark.id}), 201
    except Exception as e:
        db.rollback()
        print(f"[BOOKMARKS] ❌ Error: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/api/bookmarks/<int:bookmark_id>', methods=['DELETE'])
@login_required
def delete_bookmark(bookmark_id):
    """Remove a bookmark."""
    db = get_db()
    try:
        bookmark = db.query(BookmarkedArticle).filter(
            BookmarkedArticle.id == bookmark_id,
            BookmarkedArticle.user_id == current_user.id
        ).first()
        if not bookmark:
            return jsonify({'error': 'not_found'}), 404
        db.delete(bookmark)
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/api/bookmarks/check', methods=['POST'])
@login_required
def check_bookmarks():
    """Check which URLs are already bookmarked (batch check)."""
    data = request.get_json() or {}
    urls = data.get('urls', [])
    if not urls:
        return jsonify({'bookmarked': {}})

    db = get_db()
    try:
        existing = db.query(BookmarkedArticle.original_url, BookmarkedArticle.id).filter(
            BookmarkedArticle.user_id == current_user.id,
            BookmarkedArticle.original_url.in_(urls)
        ).all()
        bookmarked = {row.original_url: row.id for row in existing}
        return jsonify({'bookmarked': bookmarked})
    finally:
        db.close()


# =============================================================================
# DATA LIFECYCLE MANAGEMENT - Monthly reset on the 1st of every month
# =============================================================================

def _next_first_of_month():
    """Return the datetime for the 1st of the next month (00:00 UTC)."""
    now = datetime.utcnow()
    if now.month == 12:
        return datetime(now.year + 1, 1, 1)
    return datetime(now.year, now.month + 1, 1)

def check_and_run_cleanup():
    """Delete ALL articles and old monitor jobs on the 1st of every month.
    Runs on server startup — if today is the 1st, wipe everything."""
    from models import Article, MonitorJob
    now = datetime.utcnow()
    
    if now.day != 1:
        next_reset = _next_first_of_month()
        days_left = (next_reset - now).days
        print(f"[CLEANUP] ℹ️ Not the 1st — next reset in {days_left} days ({next_reset.strftime('%Y-%m-%d')})")
        return False
    
    db = get_db()
    try:
        article_count = db.query(Article).count()
        if article_count == 0:
            print(f"[CLEANUP] ℹ️ Monthly reset day but no articles to delete")
            return False
        
        db.query(Article).delete()
        
        job_count = db.query(MonitorJob).filter(
            MonitorJob.status.in_(['SUCCEEDED', 'FAILED'])
        ).delete(synchronize_session=False)
        
        db.commit()
        print(f"[CLEANUP] 🗑️ Monthly reset: deleted {article_count} articles, {job_count} old jobs")
        return True
    except Exception as e:
        print(f"[CLEANUP] ❌ Error: {e}")
        db.rollback()
        return False
    finally:
        db.close()

@app.route('/api/system/cleanup-status', methods=['GET'])
@login_required
def get_cleanup_status():
    """Get monthly reset status — days until the 1st of next month."""
    db = get_db()
    try:
        user_article_count = db.query(Article).filter(
            Article.user_id == current_user.id
        ).count()
        
        now = datetime.utcnow()
        next_reset = _next_first_of_month()
        days_remaining = (next_reset - now).days
        
        # Show warning when 3 days or less until reset
        show_warning = days_remaining <= 3 and user_article_count > 0
        
        return jsonify({
            'days_remaining': days_remaining,
            'show_warning': show_warning,
            'next_reset': next_reset.strftime('%Y-%m-%d'),
            'article_count': user_article_count
        })
    finally:
        db.close()

# =============================================================================

def auto_initialize():
    """Auto-initialize database with admin user, countries, and sources on startup.
    
    This ensures the web service has necessary data even on Render where
    shell and web service don't share the same container.
    """
    from auth_utils import hash_password
    from models import engine, DATABASE_URL
    from sqlalchemy import text
    
    print("[INIT] Running auto-initialization...")
    
    # Initialize database tables
    init_db()
    
    # Run column migrations (safe to run multiple times)
    print("[INIT] Checking column migrations...")
    try:
        with engine.connect() as conn:
            is_postgres = 'postgresql' in DATABASE_URL or 'postgres' in DATABASE_URL
            
            if is_postgres:
                # PostgreSQL: Add columns if not exist
                conn.execute(text("ALTER TABLE keywords ADD COLUMN IF NOT EXISTS translations_json TEXT"))
                conn.execute(text("ALTER TABLE keywords ADD COLUMN IF NOT EXISTS translations_updated_at TIMESTAMP"))
                conn.execute(text("ALTER TABLE articles ALTER COLUMN url TYPE VARCHAR(2000)"))
                conn.execute(text("ALTER TABLE articles ALTER COLUMN image_url TYPE VARCHAR(2000)"))
                conn.execute(text("ALTER TABLE sources ALTER COLUMN url TYPE VARCHAR(2000)"))
                # Widen daily_briefs.date_key for keyword-specific cache keys (YYYY-MM-DD:keyword)
                conn.execute(text("ALTER TABLE daily_briefs ALTER COLUMN date_key TYPE VARCHAR(100)"))
                # Exports file_data for Render compatibility
                conn.execute(text("ALTER TABLE exports ADD COLUMN IF NOT EXISTS file_data BYTEA"))
                # Migrate articles unique constraint: url-only → (url, user_id)
                result = conn.execute(text("""
                    SELECT 1 FROM pg_constraint 
                    WHERE conname = 'uq_article_url_user' AND conrelid = 'articles'::regclass
                """))
                if not result.fetchone():
                    # Drop old unique on url alone (may have different names)
                    conn.execute(text("""
                        DO $$ BEGIN
                            EXECUTE (SELECT string_agg('ALTER TABLE articles DROP CONSTRAINT ' || conname, '; ')
                                     FROM pg_constraint
                                     WHERE conrelid = 'articles'::regclass
                                       AND contype = 'u'
                                       AND array_length(conkey, 1) = 1);
                        EXCEPTION WHEN OTHERS THEN NULL;
                        END $$;
                    """))
                    conn.execute(text("DROP INDEX IF EXISTS ix_articles_url"))
                    conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_article_url_user ON articles(url, user_id)"))
                    print("[INIT] ✅ PostgreSQL articles unique constraint migrated to (url, user_id)")
                conn.commit()
                print("[INIT] ✅ PostgreSQL columns migrated")
            else:
                # SQLite: Check and add columns
                result = conn.execute(text("PRAGMA table_info(keywords)"))
                columns = [row[1] for row in result]
                if 'translations_json' not in columns:
                    conn.execute(text("ALTER TABLE keywords ADD COLUMN translations_json TEXT"))
                    conn.execute(text("ALTER TABLE keywords ADD COLUMN translations_updated_at TIMESTAMP"))
                    conn.commit()
                    print("[INIT] ✅ SQLite keywords columns migrated")
                
                # Exports file_data for Render compatibility
                result = conn.execute(text("PRAGMA table_info(exports)"))
                export_columns = [row[1] for row in result]
                if export_columns and 'file_data' not in export_columns:
                    conn.execute(text("ALTER TABLE exports ADD COLUMN file_data BLOB"))
                    conn.commit()
                    print("[INIT] ✅ SQLite exports file_data column migrated")
                
                # Migrate articles unique constraint: url-only → (url, user_id)
                # SQLite can't ALTER constraints, so we must recreate the table
                result = conn.execute(text("PRAGMA index_list(articles)"))
                indexes = [(row[1], row[2]) for row in result]
                has_new_composite = False
                needs_migration = False
                for idx_name, idx_unique in indexes:
                    if idx_unique:
                        idx_info = conn.execute(text(f"PRAGMA index_info('{idx_name}')"))
                        idx_cols = [r[2] for r in idx_info]
                        if idx_cols == ['url']:
                            needs_migration = True
                        if set(idx_cols) == {'url', 'user_id'}:
                            has_new_composite = True
                
                if needs_migration and not has_new_composite:
                    print("[INIT] Migrating articles table: url unique → (url, user_id) composite...")
                    # Get actual columns from existing table
                    col_result = conn.execute(text("PRAGMA table_info(articles)"))
                    old_cols = [row[1] for row in col_result]
                    col_list = ', '.join(old_cols)
                    
                    conn.execute(text("ALTER TABLE articles RENAME TO _articles_old_migration"))
                    # Recreate with same columns but composite unique instead of url-only unique
                    col_defs = []
                    for c in old_cols:
                        if c == 'id':
                            col_defs.append('id INTEGER PRIMARY KEY')
                        elif c == 'country':
                            col_defs.append('country VARCHAR(100) NOT NULL')
                        elif c == 'source_name':
                            col_defs.append('source_name VARCHAR(200) NOT NULL')
                        elif c == 'url':
                            col_defs.append('url VARCHAR(2000) NOT NULL')
                        elif c == 'title_original':
                            col_defs.append('title_original TEXT NOT NULL')
                        elif c == 'user_id':
                            col_defs.append('user_id INTEGER REFERENCES users(id)')
                        elif c in ('published_at', 'fetched_at', 'created_at'):
                            col_defs.append(f'{c} DATETIME')
                        elif c in ('sentiment', 'sentiment_label', 'language', 'original_language'):
                            col_defs.append(f'{c} VARCHAR(50)')
                        elif c in ('keyword', 'keyword_original'):
                            col_defs.append(f'{c} VARCHAR(200)')
                        elif c == 'sentiment_score':
                            col_defs.append(f'{c} VARCHAR(20)')
                        elif c == 'source_id':
                            col_defs.append(f'{c} INTEGER')
                        else:
                            col_defs.append(f'{c} TEXT')
                    col_defs.append('UNIQUE(url, user_id)')
                    create_sql = f"CREATE TABLE articles ({', '.join(col_defs)})"
                    conn.execute(text(create_sql))
                    conn.execute(text("CREATE INDEX IF NOT EXISTS ix_articles_user_id ON articles(user_id)"))
                    conn.execute(text(f"INSERT INTO articles ({col_list}) SELECT {col_list} FROM _articles_old_migration"))
                    conn.execute(text("DROP TABLE _articles_old_migration"))
                    conn.commit()
                    print("[INIT] ✅ SQLite articles unique constraint migrated to (url, user_id)")
                elif not has_new_composite:
                    conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_article_url_user ON articles(url, user_id)"))
                    conn.commit()
                    print("[INIT] ✅ SQLite articles composite index created")
    except Exception as e:
        print(f"[INIT] ⚠️ Column migration note: {str(e)[:100]}")

    # ── Performance indexes for aggregation queries ──────────────────
    try:
        from sqlalchemy import text as _idx_text
        with engine.connect() as conn:
            _perf_indexes = [
                ("ix_articles_country",         "articles", "country"),
                ("ix_articles_url",             "articles", "url"),
                ("ix_articles_source_name",     "articles", "source_name"),
                ("ix_articles_keyword_original", "articles", "keyword_original"),
                ("ix_articles_sentiment_label", "articles", "sentiment_label"),
                ("ix_articles_created_at",      "articles", "created_at"),
                ("ix_articles_country_url",     "articles", "country, url"),
                ("ix_articles_country_created",  "articles", "country, created_at DESC"),
            ]
            created = 0
            for idx_name, table, cols in _perf_indexes:
                try:
                    conn.execute(_idx_text(f"CREATE INDEX IF NOT EXISTS {idx_name} ON {table} ({cols})"))
                    created += 1
                except Exception:
                    pass  # Index already exists or DB doesn't support IF NOT EXISTS
            conn.commit()
            if created:
                print(f"[INIT] ✅ Ensured {created} performance indexes on articles table")
    except Exception as e:
        print(f"[INIT] ⚠️ Index creation note: {str(e)[:120]}")

    db = get_db()
    try:
        # Ensure system user exists (user_id=0 for global scheduler jobs)
        system_user = db.query(User).filter(User.id == 0).first()
        if not system_user:
            from sqlalchemy import text as sa_text
            try:
                db.execute(sa_text(
                    "INSERT INTO users (id, email, name, password_hash, role, is_active) "
                    "VALUES (0, 'system@internal', 'System', 'NOLOGIN', 'SYSTEM', false)"
                ))
                db.commit()
                print("[INIT] ✅ System user created (id=0)")
            except Exception:
                db.rollback()  # Already exists or auto-increment conflict
        
        # C1 FIX: Admin password from env var (never hardcoded)
        import secrets as _init_sec
        _admin_pw = os.environ.get('ADMIN_INIT_PASSWORD', '').strip()
        _pw_was_generated = False
        if not _admin_pw:
            _admin_pw = _init_sec.token_urlsafe(16)
            _pw_was_generated = True

        admin = db.query(User).filter(User.email == "elite@local").first()
        if not admin:
            admin = User(
                name="عبدالله الكلثمي",
                email="elite@local",
                password_hash=hash_password(_admin_pw),
                role="ADMIN",
                is_active=True,
            )
            db.add(admin)
            db.commit()
            print(f"[INIT] ✅ Admin user created: elite@local")
            if _pw_was_generated:
                print(f"[INIT] ⚠️ Generated admin password: {_admin_pw}")
                print("[INIT]    Set ADMIN_INIT_PASSWORD env var to use a fixed password.")
        else:
            # Ensure admin is active and has correct role
            admin.is_active = True
            admin.role = "ADMIN"
            if not admin.password_hash or len(admin.password_hash) < 10:
                admin.password_hash = hash_password(_admin_pw)
                if _pw_was_generated:
                    print(f"[INIT] ⚠️ Admin password was missing/invalid. New password: {_admin_pw}")
            db.commit()
            print("[INIT] ✅ Admin user verified: elite@local")
        
        # Check if we have countries, if not seed them
        country_count = db.query(Country).count()
        if country_count == 0:
            print("[INIT] Seeding countries...")
            from seed_data import seed_database
            seed_database()
            country_count = db.query(Country).count()
            print(f"[INIT] ✅ Seeded {country_count} countries")
        else:
            print(f"[INIT] ✅ Countries: {country_count}")
        
        # Check if we have sources
        source_count = db.query(Source).count()
        if source_count == 0:
            print("[INIT] ⚠️ No sources found - seeding should have added them")
        else:
            print(f"[INIT] ✅ Sources: {source_count}")
            
    except Exception as e:
        print(f"[INIT] ❌ Error during initialization: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()
    
    # Check and run monthly data reset (1st of every month)
    check_and_run_cleanup()
    
    # Auto-start global scheduler if any user has keywords
    from global_scheduler import global_scheduler
    db2 = get_db()
    try:
        keyword_count = db2.query(Keyword).filter(Keyword.enabled == True).count()
        if keyword_count > 0:
            print(f"[INIT] 🚀 Auto-starting global scheduler ({keyword_count} active keywords)")
            global_scheduler.start()
        else:
            print(f"[INIT] ℹ️ No active keywords - global scheduler will start when keywords are added")
    except Exception as e:
        print(f"[INIT] ⚠️ Could not auto-start scheduler: {e}")
    finally:
        db2.close()

    # Pre-warm map cache in background so first user doesn't wait
    def _warm_map_cache():
        try:
            _map_cache['data'] = _build_map_cache()
            _map_cache['ts'] = datetime.utcnow()
            print(f"[INIT] ✅ Map cache pre-warmed")
        except Exception as e:
            print(f"[INIT] ⚠️ Map cache warm-up failed: {e}")
    _threading.Thread(target=_warm_map_cache, daemon=True).start()


# Run auto-initialization when module loads (for Gunicorn)
auto_initialize()


if __name__ == '__main__':
    print("\n" + "="*50)
    print("🌍 عين (Ain) - News Monitor")
    print("="*50 + "\n")

    
    # Show translation service
    print("✅ Translation: Google Translate (FREE)")
    print("ℹ️  Sentiment Analysis: Disabled (future update)")
    
    print("\n")
    # Bind to Render's PORT if available; fall back to 5555 for local dev.
    import os
    port = int(os.environ.get("PORT", 5555))
    # Disable reloader to prevent interruptions during monitoring
    app.run(debug=False, use_reloader=False, host='0.0.0.0', port=port)
